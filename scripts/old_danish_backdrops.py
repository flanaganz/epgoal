#!/usr/bin/env python3
"""
danish_backdrops.py — separat sideprojekt til danske TMDb-backdrops (Mac mini / always-on)

FORMÅL
    UHF vælger selv billeder fra TMDb for programmer uden <icon>/<backdrop> i
    XML'en, men understøtter IKKE at vælge foretrukket sprog. Dette script
    slår titler op på TMDb, tjekker om der findes et ÆGTE dansk BACKDROP
    (include_image_language=da, UDEN fallback til andre sprog), og indlejrer
    det direkte i XML'en - men KUN for titler, du selv har godkendt manuelt.

KUN BACKDROPS (rettet 2026-08-07) - POSTERS BRUGES IKKE
    Tidligere injicerede scriptet BÅDE <icon> (TMDb-poster, naturligt format
    2:3/portræt) OG <backdrop> (TMDb-backdrop, naturligt format 16:9/
    landskab). Det viste sig at UHF viser <icon> i en bred 16:9-ramme, hvilket
    fik portræt-postere til at blive kraftigt beskåret/zoomet ind. Brugeren
    bruger udelukkende backdrops, så scriptet er nu renset for al
    poster/icon-håndtering:
    - resolve_danish_artwork henter STADIG poster_path fra TMDb's API-svar
      (det koster intet ekstra, samme API-kald), men kun backdrop-URL'en
      bruges til at afgøre "fundet" og til selve injektionen.
    - Der tilføjes ALDRIG et <icon>-tag i XML'en fra dette script.
    - Excel-godkendelsesfilen har ikke længere en "Dansk Poster"-kolonne
      (se export_danish_artwork_review.py).

GODKENDELSES-WORKFLOW
    1) Scriptet slår altid nye titler op og gemmer fund i
       data/danish_artwork_cache.json.
    2) Kør scripts/export_danish_artwork_review.py for at eksportere alle
       BACKDROP-fund til data/danish_artwork_review.xlsx.
    3) Åbn Excel-filen, markér "X" i kolonnen "Godkendt (X)", gem filen.
    4) Kør dette script igen - kun de X-markerede titler får deres backdrop
       indsat.

    Hvis data/danish_artwork_review.xlsx slet ikke findes endnu, injicerer
    scriptet INTET og fortæller dig at køre eksport-scriptet først.

ADSKILLELSE FRA SPORTS-SCRIPTET (enrich_epg.py) - VIGTIGT
    - Dette script rører ALDRIG data/cache.json (sports-scriptets cache).
    - Det har sin egen cache: data/danish_artwork_cache.json.
    - Det springer AUTOMATISK alle programmer over, der allerede har et
      <icon> eller <backdrop> tag (dvs. alt sport-kurateret indhold).
    - Kør ALTID dette script EFTER enrich_epg.py.

BRUG
    python3 scripts/danish_backdrops.py
    python3 scripts/danish_backdrops.py --limit 50      (test på kun 50 nye unikke titler)
    python3 scripts/danish_backdrops.py --files denmark1 (kør kun på én fil)
"""
from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
from pathlib import Path
from xml.etree import ElementTree as ET

import requests

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
CONFIG_FILE = ROOT / "config.json"

DANISH_ARTWORK_CACHE_FILE = DATA_DIR / "danish_artwork_cache.json"
DANISH_ARTWORK_REVIEW_FILE = DATA_DIR / "danish_artwork_review.xlsx"
DANISH_BACKDROPS_RUN_LOG_FILE = DATA_DIR / "danish_backdrops_run_log.json"
MAX_RUN_LOG_ENTRIES = 100

ENV_FILE = ROOT / ".env"
if ENV_FILE.exists():
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

TMDB_API_KEY = os.environ.get("TMDB_API_KEY")
TMDB_BASE = "https://api.themoviedb.org/3"
IMAGE_BASE = "https://image.tmdb.org/t/p"

MATCH_SIMILARITY_MIN = 0.55
REQUEST_SLEEP_SECONDS = 0.05

INVISIBLE_CHARS_PATTERN = re.compile(r"[\u200B\u200C\u200D\u2060\uFEFF\u00AD]")

SESSION = requests.Session()


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            print(f"⚠️  Kunne ikke læse {path}, bruger default", file=sys.stderr)
    return default


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def normalize_title(title: str) -> str:
    t = title.strip()
    t = unicodedata.normalize("NFKC", t)
    t = INVISIBLE_CHARS_PATTERN.sub(" ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip().lower()


def load_approved_keys(review_path: Path) -> set[str] | None:
    if not review_path.exists():
        return None

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl er ikke installeret - kan ikke læse godkendelsesfilen. "
              "Kør: pip install openpyxl", file=sys.stderr)
        return None

    wb = load_workbook(review_path, data_only=True)
    ws = wb.active

    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Nøgle (intern)")
        godkendt_col = headers.index("Godkendt (X)")
    except ValueError:
        print(f"⚠️  {review_path.name} mangler forventede kolonner - ingen titler godkendes denne gang.",
              file=sys.stderr)
        return set()

    approved: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        godkendt_val = row[godkendt_col].value
        if key_val and str(godkendt_val).strip().upper() == "X":
            approved.add(str(key_val).strip())
    return approved


def tmdb_search(title: str) -> tuple[str, int] | None:
    resp = SESSION.get(
        f"{TMDB_BASE}/search/multi",
        params={"api_key": TMDB_API_KEY, "query": title, "language": "da-DK", "include_adult": "false"},
        timeout=15,
    )
    resp.raise_for_status()
    results = [r for r in resp.json().get("results", []) if r.get("media_type") in ("tv", "movie")]
    if not results:
        return None

    def score(r):
        name = r.get("name") or r.get("title") or ""
        similarity = difflib.SequenceMatcher(None, name.lower(), title.lower()).ratio()
        popularity_bonus = min(r.get("popularity", 0), 50) / 50
        return similarity * 2 + popularity_bonus

    results.sort(key=score, reverse=True)
    best = results[0]
    best_name = best.get("name") or best.get("title") or ""
    similarity = difflib.SequenceMatcher(None, best_name.lower(), title.lower()).ratio()
    if similarity < MATCH_SIMILARITY_MIN:
        return None
    return best["media_type"], best["id"]


def tmdb_danish_images(media_type: str, tmdb_id: int) -> tuple[str | None, str | None]:
    """Returnerer (backdrop_path, poster_path). Poster hentes stadig fra API-
    svaret (koster intet ekstra), men bruges IKKE længere til noget - se
    modulets docstring. Beholdt for evt. fremtidig brug/debugging."""
    resp = SESSION.get(
        f"{TMDB_BASE}/{media_type}/{tmdb_id}/images",
        params={"api_key": TMDB_API_KEY, "include_image_language": "da"},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()

    def pick_best(items):
        if not items:
            return None
        items = sorted(items, key=lambda b: -b.get("vote_average", 0))
        return items[0]["file_path"]

    return pick_best(data.get("backdrops", [])), pick_best(data.get("posters", []))


def resolve_danish_artwork(raw_title: str, cache: dict, cache_max_age_days: int,
                           backdrop_size: str, poster_size: str) -> tuple[dict | None, bool]:
    """Returnerer (dict med 'backdrop' ELLER None hvis intet dansk BACKDROP
    fundet, bool om resultatet kom fra cache). 'found' afgøres UDELUKKENDE af
    backdrop - et dansk poster uden dansk backdrop tæller IKKE som fundet."""
    key = normalize_title(raw_title)

    cached = cache.get(key)
    if cached is not None and (time.time() - cached.get("ts", 0)) / 86400 < cache_max_age_days:
        if cached.get("backdrop"):
            return {"backdrop": cached.get("backdrop"), "poster": cached.get("poster")}, True
        return None, True

    backdrop_url = poster_url = None
    try:
        match = tmdb_search(raw_title)
        if match:
            media_type, tmdb_id = match
            b_path, p_path = tmdb_danish_images(media_type, tmdb_id)
            if b_path:
                backdrop_url = f"{IMAGE_BASE}/{backdrop_size}{b_path}"
            if p_path:
                poster_url = f"{IMAGE_BASE}/{poster_size}{p_path}"
    except requests.RequestException as exc:
        print(f"   TMDb-fejl for '{raw_title}': {exc}", file=sys.stderr)
        return None, False

    cache[key] = {"title": raw_title, "backdrop": backdrop_url, "poster": poster_url, "ts": time.time()}

    if backdrop_url:
        return {"backdrop": backdrop_url, "poster": poster_url}, False
    return None, False


def process_xml_file(xml_path: Path, cache: dict, cache_max_age_days: int,
                      backdrop_size: str, poster_size: str, limit: int | None,
                      titles_processed_this_run: set, approved_keys: set[str] | None,
                      all_found_keys: set[str]) -> dict:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    stats = {
        "programmes": 0, "already_had_artwork": 0, "checked": 0,
        "danish_found": 0, "danish_not_found": 0, "danish_injected": 0,
        "cache_hits": 0, "fresh_calls": 0,
    }

    resolved_this_file: dict[str, dict | None] = {}

    for programme in root.findall("programme"):
        stats["programmes"] += 1

        if programme.find("icon") is not None or programme.find("backdrop") is not None:
            stats["already_had_artwork"] += 1
            continue

        title_el = programme.find("title")
        if title_el is None or not title_el.text:
            continue
        title = title_el.text
        norm = normalize_title(title)

        if limit is not None and norm not in titles_processed_this_run and len(titles_processed_this_run) >= limit:
            continue

        if norm not in resolved_this_file:
            art, from_cache = resolve_danish_artwork(title, cache, cache_max_age_days, backdrop_size, poster_size)
            resolved_this_file[norm] = art
            titles_processed_this_run.add(norm)
            stats["checked"] += 1
            if from_cache:
                stats["cache_hits"] += 1
            else:
                stats["fresh_calls"] += 1
                time.sleep(REQUEST_SLEEP_SECONDS)
            if art:
                stats["danish_found"] += 1
                all_found_keys.add(norm)
            else:
                stats["danish_not_found"] += 1

        art = resolved_this_file[norm]
        if art and approved_keys is not None and norm in approved_keys:
            # KUN backdrop indsættes - intet <icon>/poster (se modulets docstring)
            el = ET.SubElement(programme, "backdrop")
            el.set("src", art["backdrop"])
            stats["danish_injected"] += 1

    tree.write(xml_path, encoding="utf-8", xml_declaration=True)
    return stats


def append_run_log(log_path: Path, per_file_stats: dict, grand_total: dict,
                    cache_size_before: int, cache_size_after: int,
                    approved_count: int, review_exists: bool,
                    unique_found: int, unique_pending: int) -> None:
    history = load_json(log_path, [])
    if not isinstance(history, list):
        history = []

    history.append({
        "timestamp": time.time(),
        "date_str": time.strftime("%Y-%m-%d %H:%M:%S"),
        "per_file": per_file_stats,
        "totals": grand_total,
        "cache_size_before": cache_size_before,
        "cache_size_after": cache_size_after,
        "approved_count": approved_count,
        "unique_found": unique_found,
        "unique_pending": unique_pending,
        "review_file_existed": review_exists,
    })
    history = history[-MAX_RUN_LOG_ENTRIES:]
    save_json(log_path, history)


def git_push(repo_dir: Path, commit_message: str) -> None:
    print("\n⬆️  Committer og pusher til GitHub ...")
    try:
        subprocess.run(["git", "add", "-A"], cwd=repo_dir, check=False)
        result = subprocess.run(
            ["git", "commit", "-m", commit_message], cwd=repo_dir, check=False, capture_output=True, text=True
        )
        if "nothing to commit" in (result.stdout + result.stderr).lower():
            print("   Ingen ændringer at committe.")
            return
        subprocess.run(["git", "push", "origin", "main"], cwd=repo_dir, check=False)
        print("✅ Git push forsøgt (tjek GitHub for resultat).")
    except FileNotFoundError:
        print("⚠️  git blev ikke fundet i PATH — spring commit/push over.", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(description="Tilføj GODKENDTE danske TMDb-backdrops til ikke-sport-programmer.")
    parser.add_argument("--limit", type=int, default=None,
                        help="Maks antal NYE (ikke-cachede) unikke titler at slå op.")
    parser.add_argument("--files", nargs="*", default=None,
                        help="Kør kun på specifikke kildenavne (fx denmark1 denmark3).")
    args = parser.parse_args()

    if not TMDB_API_KEY:
        sys.exit("❌ TMDB_API_KEY er ikke sat i .env - kan ikke slå danske backdrops op.")

    config = load_json(CONFIG_FILE, {})
    sources = config.get("sources", [])
    if not sources:
        sys.exit("❌ Ingen kilder defineret i config.json.")

    source_names = [s["name"] for s in sources]
    if args.files:
        source_names = [n for n in source_names if n in args.files]
        if not source_names:
            sys.exit("❌ Ingen af de angivne --files matcher kilderne i config.json.")

    backdrop_size = config.get("image", {}).get("backdrop_size", "w1280")
    poster_size = config.get("image", {}).get("poster_size", "w500")
    cache_max_age_days = config.get("cache_max_age_days", 30)
    git_cfg = config.get("git", {})

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    cache = load_json(DANISH_ARTWORK_CACHE_FILE, {})
    cache_size_before = len(cache)

    approved_keys = load_approved_keys(DANISH_ARTWORK_REVIEW_FILE)
    review_exists = DANISH_ARTWORK_REVIEW_FILE.exists()

    print("=== Danske TMDb-backdrops (separat sideprojekt) — KUN backdrops ===")
    print(f"Cache-fil: {DANISH_ARTWORK_CACHE_FILE.name} (separat fra sports-scriptets cache.json)")
    print(f"Cache indeholder {cache_size_before:,} tidligere opslag (levetid: {cache_max_age_days} dage)")

    if approved_keys is None:
        print(f"⚠️  {DANISH_ARTWORK_REVIEW_FILE.name} findes IKKE endnu.")
        print("    Der injiceres INGEN danske billeder i denne kørsel (kun opslag/cache opdateres).")
        print("    Kør 'python3 scripts/export_danish_artwork_review.py' bagefter,")
        print("    markér 'X' for de rigtige fund i Excel-filen, og kør dette script igen.")
        approved_keys = set()
    else:
        print(f"Godkendelsesfil fundet: {len(approved_keys):,} unikke titler markeret med X vil blive injiceret.")

    if args.limit:
        print(f"⚠️  TEST-TILSTAND: maks {args.limit} NYE unikke titler slås op i denne kørsel")
    print(f"Behandler filer: {', '.join(source_names)}")

    titles_processed_this_run: set = set()
    all_found_keys: set[str] = set()
    grand_total = {
        "programmes": 0, "already_had_artwork": 0, "checked": 0,
        "danish_found": 0, "danish_not_found": 0, "danish_injected": 0,
        "cache_hits": 0, "fresh_calls": 0,
    }
    per_file_stats: dict[str, dict] = {}

    for name in source_names:
        xml_path = OUTPUT_DIR / f"{name}.xml"
        if not xml_path.exists():
            print(f"\n⚠️  {xml_path} findes ikke - har du kørt enrich_epg.py først? Springer over.")
            continue

        print(f"\n📄 Behandler {xml_path.name} ...")
        stats = process_xml_file(
            xml_path, cache, cache_max_age_days, backdrop_size, poster_size,
            args.limit, titles_processed_this_run, approved_keys, all_found_keys,
        )
        save_json(DANISH_ARTWORK_CACHE_FILE, cache)

        print(f"   Programmer i alt              : {stats['programmes']:,}")
        print(f"   Sprunget over (sport)          : {stats['already_had_artwork']:,}")
        print(f"   Titler tjekket (denne fil)     : {stats['checked']:,}")
        print(f"   Dansk BACKDROP fundet (denne fil): {stats['danish_found']:,}")
        print(f"   Heraf GODKENDT og indsat i XML : {stats['danish_injected']:,} (forekomster, ikke unikke titler)")
        print(f"   Bekræftet intet dansk backdrop : {stats['danish_not_found']:,}")
        print(f"   (cache: {stats['cache_hits']:,} / friske TMDb-kald: {stats['fresh_calls']:,})")

        per_file_stats[name] = stats
        for k in grand_total:
            grand_total[k] += stats[k]

    save_json(DANISH_ARTWORK_CACHE_FILE, cache)
    cache_size_after = len(cache)

    unique_found = len(all_found_keys)
    unique_approved_and_found = len(all_found_keys & approved_keys)
    unique_pending = unique_found - unique_approved_and_found

    append_run_log(
        DANISH_BACKDROPS_RUN_LOG_FILE, per_file_stats, grand_total,
        cache_size_before, cache_size_after, len(approved_keys), review_exists,
        unique_found, unique_pending,
    )

    print("\n📊 SAMLET RAPPORT")
    print("--------------------------------")
    print(f"Programmer i alt              : {grand_total['programmes']:,}")
    print(f"Sprunget over (sport)          : {grand_total['already_had_artwork']:,}")
    print(f"Injektioner i alt (forekomster): {grand_total['danish_injected']:,} "
          "(SAMME titel kan indsættes flere gange, én gang pr. udsendelse i skemaet)")
    print()
    print(f"--- SANDE UNIKKE TAL (dedupliceret på tværs af alle {len(source_names)} filer) ---")
    print(f"Unikke titler med dansk BACKDROP fundet : {unique_found:,}")
    print(f"  - heraf GODKENDT (X i Excel)           : {unique_approved_and_found:,}")
    print(f"  - heraf AFVENTER stadig godkendelse    : {unique_pending:,}")
    print(f"Cache voksede fra {cache_size_before:,} til {cache_size_after:,} unikke titler i alt")
    if unique_pending > 0:
        print(f"💡 Åbn {DANISH_ARTWORK_REVIEW_FILE.name} og markér flere 'X' for at godkende de resterende "
              f"{unique_pending:,} fund.")
    print("--------------------------------")

    if git_cfg.get("enabled", True):
        prefix = "Auto-sync EPG (danske backdrops)"
        git_push(ROOT, f"{prefix} {time.strftime('%Y-%m-%d %H:%M:%S')}")

    print("\n=== Færdig. ===")


if __name__ == "__main__":
    main()
