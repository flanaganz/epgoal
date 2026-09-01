#!/usr/bin/env python3
"""
export_danish_artwork_review.py — eksporterer danske TMDb BACKDROP-fund til Excel til manuel godkendelse

FORMÅL
    Læser data/danish_artwork_cache.json (fra danish_backdrops.py) og skriver
    alle titler MED et fundet dansk BACKDROP til
    data/danish_artwork_review.xlsx, så du kan gennemgå dem og markere "X" i
    kolonnen "Godkendt (X)" for de fund, der er korrekte og skal bruges.

    KUN BACKDROPS (rettet 2026-08-07): Der er ingen "Dansk Poster"-kolonne
    længere - postere bruges ikke (UHF viste dem forkert beskåret/zoomet, da
    de er portræt-format vist i en 16:9-ramme). Titler der KUN har et dansk
    poster, men intet dansk backdrop, optræder derfor slet ikke i denne fil.

    Kør dette script igen, når du har fundet flere danske backdrops over tid.
    Eksisterende X-markeringer og noter BEVARES automatisk for titler, der
    allerede findes i filen. Kun NYE titler tilføjes som friske, tomme rækker.

RETTET (2026-09-01): ENSARTET MED sport_artwork_review.xlsx - "IGNORER (X)"-KOLONNE
    Tidligere skulle du skrive fri tekst (fx "Ignore"/"Ignorer") i "Note" for
    at markere en titel som forkert/uønsket - en ren tekstkonvention uden
    nogen kolonnevalidering, i modsætning til sport_artwork_review.xlsx, som
    allerede har en dedikeret "Ignorer (X)"-afkrydsningskolonne. Det er nu
    ensartet: denne fil har fået samme "Ignorer (X)"-kolonne (med dropdown),
    lige efter "Godkendt (X)".

    BAGUDKOMPATIBELT: findes "Ignorer (X)"-kolonnen IKKE i en ældre fil,
    migreres eksisterende rækker automatisk - en Note der (efter trim, uden
    hensyn til store/små bogstaver) STARTER MED "ignor" (dækker både
    "Ignore" og "Ignorer") sættes til Ignorer (X) = "X", og selve Note-feltet
    ryddes (teksten var jo kun en markør, ikke en rigtig kommentar). Alle
    andre, "rigtige" kommentarer i Note (der ikke starter med "ignor")
    bevares uændret.

    Ligesom i sport_artwork_review.xlsx SKJULES/UDELADES ignorerede rækker
    IKKE fra arket - de vises fortsat, blot med "Ignorer (X)" allerede sat
    til "X", så du altid kan se og om nødvendigt fortryde et valg. Brug
    ExcEls indbyggede autofilter (kolonneoverskriften har et filter-ikon)
    til at skjule dem, hvis du vil have et renere overblik.

    danish_backdrops.py og generate_stats_report.py respekterer nu også
    denne kolonne (se de respektive filer for detaljer).

WORKFLOW
    1) python3 scripts/danish_backdrops.py           (finder danske backdrops)
    2) python3 scripts/export_danish_artwork_review.py  (eksporter til Excel)
    3) Åbn data/danish_artwork_review.xlsx, markér "X", gem filen
    4) python3 scripts/danish_backdrops.py           (injicerer de godkendte)

BRUG
    python3 scripts/export_danish_artwork_review.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

DANISH_ARTWORK_CACHE_FILE = DATA_DIR / "danish_artwork_cache.json"
DANISH_ARTWORK_REVIEW_FILE = DATA_DIR / "danish_artwork_review.xlsx"

HEADERS = ["Nøgle (intern)", "Titel", "Dansk Backdrop", "Godkendt (X)", "Ignorer (X)", "Note"]

# Bruges KUN til bagudkompatibel migrering af ældre filer uden "Ignorer (X)"
# -kolonnen, hvor "Ignorer"/"Ignore" blev skrevet som fri tekst i Note.
IGNORE_NOTE_PREFIX = "ignor"  # dækker case-insensitive både "ignore" og "ignorer"


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def is_checked(value) -> bool:
    """True når en Excel-statuskolonne indeholder X (case-insensitive)."""
    return str(value or "").strip().upper() == "X"


def is_legacy_ignore_note(note_value) -> bool:
    """True hvis en GAMMEL Note-værdi (fra før 'Ignorer (X)'-kolonnen fandtes)
    var en ren ignorer-markør, dvs. starter med 'ignor' (dækker 'Ignore' og
    'Ignorer', uanset store/små bogstaver)."""
    return bool(note_value) and str(note_value).strip().lower().startswith(IGNORE_NOTE_PREFIX)


def load_existing_review(path: Path) -> dict[str, dict]:
    """Indlæs tidligere godkendelser/ignoreringer/noter med bagudkompatibilitet.

    Den aktuelle model bruger kolonnerne Godkendt (X), Ignorer (X) og Note.
    Ældre ark uden Ignorer-kolonnen understøttes også: en Note der starter
    med 'ignor' (fx det tidligere brugte 'Ignore') migreres automatisk til
    ignorer='X', og selve Note-feltet ryddes (se docstring øverst i filen)."""
    existing: dict[str, dict] = {}
    if not path.exists():
        return existing

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Nøgle (intern)")
        godkendt_col = headers.index("Godkendt (X)")
    except ValueError:
        print(f"⚠️  Eksisterende {path.name} har uventet format - starter forfra uden at bevare markeringer.",
              file=sys.stderr)
        return existing

    note_col = headers.index("Note") if "Note" in headers else None
    ignore_col = headers.index("Ignorer (X)") if "Ignorer (X)" in headers else None

    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        if not key_val:
            continue

        godkendt_val = row[godkendt_col].value
        note_val = row[note_col].value if note_col is not None else None
        ignore_val = row[ignore_col].value if ignore_col is not None else None

        # Migrer den gamle note-baserede ignorer-markør ("Ignore"/"Ignorer"
        # skrevet som fri tekst), hvis filen ikke allerede har en rigtig
        # Ignorer (X)-kolonne.
        if ignore_col is None and is_legacy_ignore_note(note_val):
            ignore_val = "X"
            note_val = None

        existing[str(key_val).strip()] = {
            "godkendt": godkendt_val,
            "ignorer": ignore_val,
            "note": note_val,
        }
    return existing


def main() -> None:
    cache = load_json(DANISH_ARTWORK_CACHE_FILE, {})
    if not cache:
        sys.exit(f"❌ {DANISH_ARTWORK_CACHE_FILE} findes ikke eller er tom - kør danish_backdrops.py først.")

    # KUN titler med et fundet BACKDROP - postere ignoreres helt (se docstring)
    candidates = {
        key: entry for key, entry in cache.items()
        if entry.get("backdrop")
    }

    if not candidates:
        sys.exit("❌ Ingen danske backdrops fundet endnu i cachen - intet at eksportere.")

    existing_review = load_existing_review(DANISH_ARTWORK_REVIEW_FILE)
    new_count = sum(1 for key in candidates if key not in existing_review)
    preserved_count = sum(1 for key in candidates if key in existing_review and existing_review[key].get("godkendt"))
    ignored_count = sum(1 for key in candidates if key in existing_review and is_checked(existing_review[key].get("ignorer")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Danske backdrops"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="2E5395")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(ord('A') + len(HEADERS) - 1)}1"

    sorted_keys = sorted(candidates.keys(), key=lambda k: candidates[k].get("title", k).lower())

    for row_idx, key in enumerate(sorted_keys, start=2):
        entry = candidates[key]
        title = entry.get("title") or key.title()
        backdrop_url = entry.get("backdrop")
        prior = existing_review.get(key, {})

        ws.cell(row=row_idx, column=1, value=key)
        ws.cell(row=row_idx, column=2, value=title)

        if backdrop_url:
            c = ws.cell(row=row_idx, column=3, value="Åbn backdrop")
            c.hyperlink = backdrop_url
            c.style = "Hyperlink"

        godkendt_val = prior.get("godkendt")
        if godkendt_val:
            ws.cell(row=row_idx, column=4, value=godkendt_val)

        if is_checked(prior.get("ignorer")):
            ws.cell(row=row_idx, column=5, value="X")

        note_val = prior.get("note")
        if note_val:
            ws.cell(row=row_idx, column=6, value=note_val)

    last_row = len(sorted_keys) + 1

    dv_godkendt = DataValidation(type="list", formula1='"X"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_godkendt)
    dv_godkendt.add(f"D2:D{last_row}")

    dv_ignorer = DataValidation(type="list", formula1='"X"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv_ignorer)
    dv_ignorer.add(f"E2:E{last_row}")

    widths = {"A": 32, "B": 45, "C": 16, "D": 14, "E": 14, "F": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    guide = wb.create_sheet("Vejledning")
    instructions = [
        ("Sådan bruger du denne fil", ""),
        ("", ""),
        ("Kolonne D - Godkendt (X)", "Markér 'X' for de fund der er KORREKTE og skal bruges som backdrop."),
        ("Kolonne E - Ignorer (X)", "Markér 'X' for titler hvor det danske TMDb-fund er FORKERT eller ikke skal bruges. "
         "Ignorerede rækker injiceres ALDRIG, selv hvis 'Godkendt (X)' også skulle være markeret ved en fejl."),
        ("Kolonne F - Note", "Din egen kommentar - bruges ikke til at afgøre noget automatisk. Skriv fx hvorfor et fund er forkert."),
        ("", ""),
        ("Vigtigt", "Kør danish_backdrops.py EFTER du har udfyldt og gemt denne fil, for at injicere de godkendte fund."),
        ("Vigtigt", "Kør export_danish_artwork_review.py igen når som helst - dine X-markeringer og noter BEVARES automatisk."),
        ("Vigtigt", "Rækker fjernes ALDRIG automatisk fra arket, heller ikke når de er markeret 'Ignorer (X)' - "
         "brug Excels autofilter (pilen i kolonneoverskriften) for at skjule dem, hvis du ønsker et renere overblik."),
    ]
    for r, (label, text) in enumerate(instructions, start=1):
        c1 = guide.cell(row=r, column=1, value=label)
        c1.font = Font(bold=(r == 1))
        guide.cell(row=r, column=2, value=text)
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 100

    wb.save(DANISH_ARTWORK_REVIEW_FILE)

    print("=== Danske backdrops eksporteret til Excel (kun backdrops, ingen postere) ===")
    print(f"Fil: {DANISH_ARTWORK_REVIEW_FILE}")
    print(f"Titler i alt (med fundet dansk backdrop): {len(candidates):,}")
    print(f"  - Nye rækker tilføjet denne gang      : {new_count:,}")
    print(f"  - Eksisterende X-markeringer bevaret  : {preserved_count:,}")
    print(f"  - Rækker markeret Ignorer (X)         : {ignored_count:,}")
    print()
    print("Åbn filen, markér 'X' i kolonnen 'Godkendt (X)' for de rigtige fund, gem filen,")
    print("og kør derefter 'python3 scripts/danish_backdrops.py' igen for at injicere dem.")


if __name__ == "__main__":
    main()
