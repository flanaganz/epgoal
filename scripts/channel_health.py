#!/usr/bin/env python3
"""
channel_health.py — scanner de FÆRDIGE output/denmarkX.xml-filer og måler
reel artwork-dækning pr. kanal.

NYT (2026-08-12): AUTOMATISK KANAL-SAMMENLÆGNING (dedup)
    Samme fysiske kanal optræder ofte under FLERE forskellige kanal-ID'er på
    tværs af de 6 Open-EPG-kilder - fx pga. HD/FHD-mærker, "Denmark (DK,DA)"-
    suffiks, eller små stave-/mellemrumsforskelle ("TV 2 Fri.dk" vs
    "TV2 fri.dk" vs "TV 2 Fri HD (D) (T).dk"). Dette script beregner nu en
    normaliseret "gruppe-nøgle" for hver kanal og lægger automatisk alle
    varianter af samme kanal sammen til ÉN række, i stedet for at vise den
    samme kanal flere gange med forskellige (ufuldstændige) tal.

    Normaliserings-reglerne (se compute_group_key()):
    - Fjerner ".dk"-suffiks, alt i parenteser (fx "(D) (T)", "(DK,DA)").
    - Fjerner "pynte-ord": hd, fhd, denmark, danmark, channel.
    - Fjerner apostroffer/kommaer/punktummer/bindestreger.
    - Fjerner ALT mellemrum til sidst - det gør at "TV 2 Fri", "TV2 Fri" og
      "TV2Fri" alle bliver til samme nøgle "tv2fri".
    - Tal og bogstaver der reelt ADSKILLER kanaler (fx "X" i "Sport X" vs
      "Sport", eller "+" i "TV3+" vs "TV3 Max") bevares bevidst, så disse
      IKKE bliver fejlagtigt slået sammen.

    Nogle kanaler kan IKKE genkendes som ens ud fra navnet alene (fx
    "Nat Geo Wild" vs "National Geographic Wild", eller "ID" vs
    "Investigation Discovery"). Til disse findes en lille indbygget
    alias-tabel (DEFAULT_ALIASES) samt mulighed for at tilføje flere i
    data/channel_group_aliases.json (valgfri fil, format: {"kanal-id eller
    display-name i små bogstaver": "ønsket gruppe-nøgle"}).

NYT (2026-08-14): KANAL-PRIORITERING ("Følg (X)")
    "Top 10 kanaler med flest MANGLENDE artwork"-listen i konsol-outputtet
    respekterer nu data/channel_priority.xlsx (samme fil som bruges af
    generate_stats_report.py) - hvis den findes, vises KUN kanaler markeret
    "X" i kolonnen "Følg (X)". Kanaler du bevidst har valgt at ignorere
    (fx DK4, France 24, div. udenlandske nyhedskanaler) fylder derfor ikke
    unødigt i "problemer der skal løses"-listen. Findes filen ikke, eller
    er ingen kanaler markeret, vises ALLE kanaler uændret (bagudkompatibelt).

NYT (2026-08-13): STATISTIK-KØRSEL VIA --stats
    Kør med flaget "--stats" for at scanne output_stats/ i stedet for
    output/, og gemme resultatet i data/channel_health_stats.json i stedet
    for data/channel_health.json. Bruges til at vurdere POTENTIEL
    artwork-dækning (fx med enrich_non_sport_with_tmdb=true), uden at røre
    produktions-snapshottet.

Skriver:
    data/channel_health.json ELLER data/channel_health_stats.json - med
    BÅDE:
      - "channels": rå data pr. oprindeligt kanal-ID (bagudkompatibelt/debug)
      - "groups": sammenlagt data pr. gruppe-nøgle (bruges af rapporten)

BRUG
    python3 scripts/channel_health.py
    python3 scripts/channel_health.py --stats
"""
from __future__ import annotations

import json
import re
import sys
import time
from pathlib import Path
from xml.etree import ElementTree as ET

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
CONFIG_FILE = ROOT / "config.json"

SUFFIX = "_stats" if "--stats" in sys.argv else ""
OUTPUT_DIR = ROOT / f"output{SUFFIX}"
CHANNEL_HEALTH_FILE = DATA_DIR / f"channel_health{SUFFIX}.json"

CHANNEL_GROUP_ALIASES_FILE = DATA_DIR / "channel_group_aliases.json"
CHANNEL_PRIORITY_FILE = DATA_DIR / "channel_priority.xlsx"

MAX_MISSING_PRINTED = 10

DECORATION_WORDS = {"hd", "fhd", "denmark", "danmark", "channel"}

# Kendte tilfælde hvor navnet alene ikke afslører at det er samme kanal.
# Nøgle = kanal-ID ELLER display-name i små bogstaver, trimmet.
# Værdi = den gruppe-nøgle det skal tilknyttes (skal matche den automatisk
# beregnede nøgle for "søster"-kanalen, se compute_group_key()).
DEFAULT_ALIASES = {
    "nat geo wild.dk": "nationalgeographicwild",
    "id.dk": "investigationdiscovery",
    "id investigation discovery (d) (t).dk": "investigationdiscovery",
}


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def load_aliases() -> dict[str, str]:
    aliases = dict(DEFAULT_ALIASES)
    extra = load_json(CHANNEL_GROUP_ALIASES_FILE, {})
    if isinstance(extra, dict):
        aliases.update({k.strip().lower(): v.strip() for k, v in extra.items()})
    return aliases


def load_channel_priority(path: Path) -> set[str] | None:
    """Returnerer sæt af GRUPPE-nøgler markeret 'X' i channel_priority.xlsx,
    eller None hvis filen ikke findes (= ingen filtrering, vis alle kanaler).
    Samme fil/format som bruges af generate_stats_report.py."""
    if not path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl er ikke installeret - kan ikke læse kanal-prioritering, viser alle kanaler.",
              file=sys.stderr)
        return None

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Kanal-prioritering"] if "Kanal-prioritering" in wb.sheetnames else wb.active
    except Exception as exc:
        print(f"⚠️  Kunne ikke åbne {path.name}: {exc} - viser alle kanaler.", file=sys.stderr)
        return None

    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Gruppe-nøgle (intern)")
        follow_col = headers.index("Følg (X)")
    except ValueError:
        print(f"⚠️  {path.name} mangler forventede kolonner - viser alle kanaler.", file=sys.stderr)
        return None

    followed: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        follow_val = row[follow_col].value
        if key_val and follow_val and str(follow_val).strip().upper() == "X":
            followed.add(str(key_val).strip())
    return followed


def compute_group_key(channel_id: str, display_name: str, aliases: dict[str, str]) -> str:
    text = display_name or channel_id
    alias = aliases.get(channel_id.strip().lower()) or aliases.get(text.strip().lower())
    if alias:
        return alias

    t = text.lower()
    if t.endswith(".dk"):
        t = t[:-3]
    t = re.sub(r"\([^)]*\)", " ", t)  # fjern parentes-indhold helt
    tokens = re.split(r"\s+", t.strip())
    tokens = [tok for tok in tokens if tok.strip(",.-") not in DECORATION_WORDS and tok.strip(",.-") != ""]
    t = " ".join(tokens)
    t = t.replace("'", "").replace(",", "").replace(".", "").replace("-", "")
    t = re.sub(r"\s+", "", t)
    return t or text.lower()


def scan_file(xml_path: Path, channel_stats: dict[str, dict]) -> int:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    display_names: dict[str, str] = {}
    for ch in root.findall("channel"):
        cid = ch.get("id", "")
        dn_el = ch.find("display-name")
        display_names[cid] = dn_el.text.strip() if dn_el is not None and dn_el.text else cid

    count = 0
    for programme in root.findall("programme"):
        count += 1
        chan_id = programme.get("channel", "")
        if chan_id not in channel_stats:
            channel_stats[chan_id] = {
                "display_name": display_names.get(chan_id, chan_id),
                "programmes": 0,
                "with_artwork": 0,
                "with_desc": 0,
            }
        stats = channel_stats[chan_id]
        stats["programmes"] += 1

        has_artwork = programme.find("icon") is not None or programme.find("backdrop") is not None
        if has_artwork:
            stats["with_artwork"] += 1

        desc_el = programme.find("desc")
        if desc_el is not None and desc_el.text and desc_el.text.strip():
            stats["with_desc"] += 1

    return count


def build_groups(channel_stats: dict[str, dict], aliases: dict[str, str]) -> dict[str, dict]:
    """Slår kanal-varianter sammen efter gruppe-nøgle. Kanonisk display-name
    vælges som den KORTESTE af de sammenlagte navne (typisk den "reneste",
    fx "TV3 Max.dk" fremfor "TV3 MAX HD (D) (T).dk")."""
    groups: dict[str, dict] = {}
    for cid, s in channel_stats.items():
        key = compute_group_key(cid, s["display_name"], aliases)
        if key not in groups:
            groups[key] = {
                "display_name": s["display_name"],
                "member_channel_ids": [],
                "programmes": 0,
                "with_artwork": 0,
                "with_desc": 0,
            }
        g = groups[key]
        g["member_channel_ids"].append(cid)
        g["programmes"] += s["programmes"]
        g["with_artwork"] += s["with_artwork"]
        g["with_desc"] += s["with_desc"]
        if len(s["display_name"]) < len(g["display_name"]):
            g["display_name"] = s["display_name"]
    return groups


def main() -> None:
    config = load_json(CONFIG_FILE, {})
    sources = config.get("sources", [])
    source_names = [s["name"] for s in sources] if sources else None

    if source_names is None:
        xml_files = sorted(OUTPUT_DIR.glob("denmark*.xml"))
    else:
        xml_files = [OUTPUT_DIR / f"{name}.xml" for name in source_names]
        xml_files = [p for p in xml_files if p.exists()]

    if not xml_files:
        sys.exit(f"❌ Ingen denmarkX.xml-filer fundet i {OUTPUT_DIR} - kør enrich_epg.py først.")

    print("=== Kanal-sundhedstjek (scanner FÆRDIGE output-filer) ===")
    print(f"Kilde-mappe: {OUTPUT_DIR}")
    print(f"Behandler {len(xml_files)} fil(er): {', '.join(p.name for p in xml_files)}")

    channel_stats: dict[str, dict] = {}
    total_programmes = 0
    for xml_path in xml_files:
        count = scan_file(xml_path, channel_stats)
        total_programmes += count
        print(f"   {xml_path.name}: {count:,} programmer")

    aliases = load_aliases()
    groups = build_groups(channel_stats, aliases)

    total_with_artwork = sum(c["with_artwork"] for c in channel_stats.values())
    overall_pct = (total_with_artwork / total_programmes * 100) if total_programmes else 0

    snapshot = {
        "timestamp": time.time(),
        "date_str": time.strftime("%Y-%m-%d %H:%M:%S"),
        "total_programmes": total_programmes,
        "total_with_artwork": total_with_artwork,
        "overall_artwork_pct": round(overall_pct, 2),
        "channels": channel_stats,   # rå data (bagudkompatibelt/debug)
        "groups": groups,            # sammenlagt data (bruges af rapporten)
    }
    save_json(CHANNEL_HEALTH_FILE, snapshot)

    print(f"\n📊 SAMLET: {total_programmes:,} programmer")
    print(f"   Rå kanal-ID'er fundet : {len(channel_stats):,}")
    print(f"   Efter sammenlægning   : {len(groups):,} unikke kanaler")
    print(f"   Artwork-dækning i alt : {total_with_artwork:,} ({overall_pct:.1f}%)")

    # --- Kanal-prioritering: filtrér "værste"-listen til kun fulgte kanaler ---
    priority = load_channel_priority(CHANNEL_PRIORITY_FILE)
    if priority is not None:
        candidate_groups = {k: v for k, v in groups.items() if k in priority}
        print(f"\n📌 Kanal-prioritering aktiv: viser kun {len(candidate_groups):,} af {len(groups):,} "
              f"kanaler markeret 'Følg (X)' i {CHANNEL_PRIORITY_FILE.name}.")
        list_label = f"Top {MAX_MISSING_PRINTED} FULGTE kanaler med flest MANGLENDE artwork"
    else:
        candidate_groups = groups
        print(f"\nℹ️  {CHANNEL_PRIORITY_FILE.name} findes ikke - viser alle kanaler "
              "(opret filen med scripts/export_channel_priority.py for at filtrere).")
        list_label = f"Top {MAX_MISSING_PRINTED} (sammenlagte) kanaler med flest MANGLENDE artwork"

    worst = sorted(candidate_groups.items(), key=lambda kv: kv[1]["programmes"] - kv[1]["with_artwork"],
                    reverse=True)[:MAX_MISSING_PRINTED]
    worst = [(key, s) for key, s in worst if s["programmes"] - s["with_artwork"] > 0]

    if worst:
        print(f"\n   {list_label}:")
        for key, s in worst:
            missing = s["programmes"] - s["with_artwork"]
            pct = (s["with_artwork"] / s["programmes"] * 100) if s["programmes"] else 0
            n_members = len(s["member_channel_ids"])
            print(f"     - {s['display_name']:35s} {missing:>5,} mangler  ({pct:.1f}% dækning, "
                  f"{s['programmes']:,} i alt, {n_members} kanal-ID(er) sammenlagt)")
    elif priority is not None:
        print(f"\n   🎉 Alle fulgte kanaler har 100% artwork-dækning!")

    print(f"\nGemt: {CHANNEL_HEALTH_FILE}")
    if not CHANNEL_GROUP_ALIASES_FILE.exists():
        print(f"💡 Tip: opret {CHANNEL_GROUP_ALIASES_FILE.name} hvis du opdager kanaler der IKKE blev "
              "automatisk sammenlagt, men reelt er den samme (fx forkortelser scriptet ikke kan gætte).")
    if priority is None:
        print(f"💡 Tip: opret {CHANNEL_PRIORITY_FILE.name} med scripts/export_channel_priority.py for kun "
              "at se dine egne prioriterede kanaler i 'manglende artwork'-listen ovenfor.")
    print("=== Færdig. ===")


if __name__ == "__main__":
    main()
