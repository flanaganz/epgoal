#!/usr/bin/env python3
"""
export_channel_priority.py — eksporterer alle SAMMENLAGTE kanaler (fra
channel_health.py's "groups") til Excel, så du kan markere hvilke kanaler
der skal FOKUSERES på i statistikrapportens "Kanal-sundhed"-sektion.

NYT (2026-08-12): én række pr. FYSISK kanal, ikke pr. rå kanal-ID. Kolonnen
"Sammenlagte kanal-ID'er" viser transparent hvilke oprindelige kanal-varianter
(fx "TV 2 Fri.dk", "TV2 fri.dk", "TV 2 Fri HD (D) (T).dk") der er lagt sammen
til denne ene række - så du kan verificere at sammenlægningen er korrekt.

Kanaler UDEN "X" i "Følg (X)" ignoreres automatisk af generate_stats_report.py.

WORKFLOW
    1) python3 scripts/channel_health.py           (scanner + sammenlægger)
    2) python3 scripts/export_channel_priority.py  (eksporter til Excel)
    3) Åbn data/channel_priority.xlsx, markér "X" for kanaler du vil følge
    4) python3 scripts/generate_stats_report.py    (rapporten respekterer dit valg)

    Kør channel_health.py + export_channel_priority.py igen når som helst -
    dine eksisterende X-markeringer BEVARES automatisk (nu matchet på
    gruppe-nøgle, som er stabil selv hvis nye kanal-ID-varianter dukker op).

BRUG
    python3 scripts/export_channel_priority.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

CHANNEL_HEALTH_FILE = DATA_DIR / "channel_health.json"
CHANNEL_PRIORITY_FILE = DATA_DIR / "channel_priority.xlsx"

HEADERS = ["Kanal", "Gruppe-nøgle (intern)", "Sammenlagte kanal-ID'er", "Programmer", "Artwork %", "Følg (X)"]


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def load_existing_priority(path: Path, channel_id_to_group: dict[str, str]) -> dict[str, str]:
    """Returnerer {gruppe-nøgle: 'X' eller tom}.

    Understøtter MIGRERING fra den ældre fil-version (før kanal-sammenlægning),
    som havde kolonnen "Kanal-ID (intern)" i stedet for "Gruppe-nøgle (intern)".
    Hvis den gamle version genkendes, oversættes dine tidligere X-markeringer
    (sat pr. rå kanal-ID) automatisk til de nye, sammenlagte gruppe-nøgler via
    channel_id_to_group - så dit tidligere arbejde IKKE går tabt."""
    existing: dict[str, str] = {}
    if not path.exists():
        return existing

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    if "Gruppe-nøgle (intern)" in headers:
        key_col = headers.index("Gruppe-nøgle (intern)")
        follow_col = headers.index("Følg (X)")
        for row in ws.iter_rows(min_row=2):
            key_val = row[key_col].value
            if not key_val:
                continue
            follow_val = row[follow_col].value
            existing[str(key_val).strip()] = str(follow_val).strip() if follow_val else ""
        return existing

    if "Kanal-ID (intern)" in headers:
        print("ℹ️  Ældre fil-version fundet (før kanal-sammenlægning) - migrerer dine X-markeringer "
              "til de nye, sammenlagte kanal-grupper ...")
        cid_col = headers.index("Kanal-ID (intern)")
        follow_col = headers.index("Følg (X)")
        migrated = 0
        for row in ws.iter_rows(min_row=2):
            cid_val = row[cid_col].value
            follow_val = row[follow_col].value
            if not cid_val or not follow_val or str(follow_val).strip().upper() != "X":
                continue
            group_key = channel_id_to_group.get(str(cid_val).strip())
            if group_key:
                if existing.get(group_key) != "X":
                    migrated += 1
                existing[group_key] = "X"
        print(f"   Migreret {migrated:,} tidligere X-markeringer til nye gruppe-nøgler.")
        return existing

    print(f"⚠️  Eksisterende {path.name} har uventet format - starter forfra uden at bevare markeringer.",
          file=sys.stderr)
    return existing


def main() -> None:
    health = load_json(CHANNEL_HEALTH_FILE, None)
    if not health or not health.get("groups"):
        sys.exit(f"❌ {CHANNEL_HEALTH_FILE} findes ikke eller mangler 'groups' - kør channel_health.py først "
                 "(evt. igen, hvis du kører en ældre version).")

    groups = health["groups"]

    # Bygges KUN til brug for migrering fra den gamle fil-version (rå kanal-ID -> gruppe-nøgle).
    channel_id_to_group: dict[str, str] = {}
    for group_key, s in groups.items():
        for cid in s.get("member_channel_ids", []):
            channel_id_to_group[cid] = group_key

    existing = load_existing_priority(CHANNEL_PRIORITY_FILE, channel_id_to_group)
    new_count = sum(1 for key in groups if key not in existing)
    preserved_count = sum(1 for key in groups if existing.get(key) == "X")

    wb = Workbook()
    ws = wb.active
    ws.title = "Kanal-prioritering"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="0f766e")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:F1"

    sorted_groups = sorted(groups.items(), key=lambda kv: kv[1]["programmes"], reverse=True)

    for row_idx, (key, s) in enumerate(sorted_groups, start=2):
        prog = s["programmes"]
        pct = (s["with_artwork"] / prog * 100) if prog else 0
        members = ", ".join(sorted(s["member_channel_ids"]))
        ws.cell(row=row_idx, column=1, value=s["display_name"])
        ws.cell(row=row_idx, column=2, value=key)
        ws.cell(row=row_idx, column=3, value=members)
        ws.cell(row=row_idx, column=4, value=prog)
        ws.cell(row=row_idx, column=5, value=round(pct, 1))
        prior = existing.get(key, "")
        if prior:
            ws.cell(row=row_idx, column=6, value=prior)

    last_row = len(sorted_groups) + 1
    dv = DataValidation(type="list", formula1='"X"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"F2:F{last_row}")

    widths = {"A": 32, "B": 26, "C": 60, "D": 14, "E": 12, "F": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Vejledning")
    instructions = [
        ("Sådan bruger du denne fil", ""),
        ("", ""),
        ("Kolonne A - Kanal", "Det kanoniske (korteste, reneste) navn for denne fysiske kanal."),
        ("Kolonne C - Sammenlagte kanal-ID'er", "Viser ALLE de oprindelige kanal-ID-varianter (fx HD/FHD/Denmark-udgaver), der automatisk er lagt sammen til denne ene række. Tjek denne kolonne hvis noget ser forkert sammenlagt ud."),
        ("Kolonne F - Følg (X)", "Markér 'X' for de kanaler du VIL følge i statistikrapporten. Resten ignoreres automatisk."),
        ("", ""),
        ("Vigtigt", "Hvis to kanaler REELT er den samme, men IKKE er blevet sammenlagt automatisk (fx en forkortelse scriptet ikke genkendte), kan du tilføje en alias i data/channel_group_aliases.json - se channel_health.py for format og eksempler."),
        ("Vigtigt", "Kør channel_health.py + export_channel_priority.py igen når nye kanaler dukker op - dine X-markeringer bevares automatisk (matchet på Gruppe-nøgle, kolonne B)."),
        ("Vigtigt", "Kør KUN generate_stats_report.py for at se opdateret statistik med dit eksisterende valg - du behøver IKKE køre export_channel_priority.py hver gang."),
    ]
    for row_idx, (label, text) in enumerate(instructions, start=1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c1.font = Font(bold=(row_idx == 1))
        ws2.cell(row=row_idx, column=2, value=text)
    ws2.column_dimensions['A'].width = 26
    ws2.column_dimensions['B'].width = 100

    wb.save(CHANNEL_PRIORITY_FILE)

    print("=== Kanal-prioriteringsliste eksporteret (sammenlagte kanaler) ===")
    print(f"Fil: {CHANNEL_PRIORITY_FILE}")
    print(f"Unikke (sammenlagte) kanaler: {len(sorted_groups):,}")
    print(f"  - Nye kanaler tilføjet denne gang   : {new_count:,}")
    print(f"  - Eksisterende X-markeringer bevaret: {preserved_count:,}")
    print()
    print("Åbn filen, markér 'X' for de kanaler du vil følge, gem filen,")
    print("og kør derefter 'python3 scripts/generate_stats_report.py' igen.")


if __name__ == "__main__":
    main()
