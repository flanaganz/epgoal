#!/usr/bin/env python3
"""
import_sport_review.py — læser dine udfyldte billed-valg fra
data/sport_artwork_review.xlsx og skriver dem ind i
data/sport_program_overrides.json, så enrich_epg.py bruger dem ved næste
kørsel.

FORMÅL
    Modstykke til export_sport_review.py. Dette script rører KUN de
    titel-nøgler, du selv har udfyldt et billede for i Excel-arket (kolonne
    "Vælg billede") - alt andet i sport_program_overrides.json bevares
    fuldstændig uændret (fx dine eksisterende overrides for Champions
    League, Ligue 1, Sporten osv. fra tidligere).

    Rækker uden noget valgt i "Vælg billede" springes stiltiende over - de
    er endnu ikke taget stilling til, og påvirker derfor ikke filen.

    Titel-nøglen normaliseres på PRÆCIS samme måde som SportMatcher i
    enrich_epg.py forventer (små bogstaver, trimmet) - se normalize_title().

WORKFLOW
    1) Udfyld og gem data/sport_artwork_review.xlsx (via export_sport_review.py)
    2) python3 scripts/import_sport_review.py
    3) python3 scripts/enrich_epg.py     (kør igen for at se de nye billeder)

BRUG
    python3 scripts/import_sport_review.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

REVIEW_FILE = DATA_DIR / "sport_artwork_review.xlsx"
SPORT_PROGRAM_OVERRIDES_FILE = DATA_DIR / "sport_program_overrides.json"


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def normalize_title(title: str) -> str:
    """Samme normalisering som enrich_epg.py's SportMatcher forventer som
    nøgle i sport_program_overrides.json: små bogstaver, trimmet, enkelt
    mellemrum."""
    import re
    import unicodedata
    t = title.strip()
    t = unicodedata.normalize("NFKC", t)
    t = re.sub(r"[\u200B\u200C\u200D\u2060\uFEFF\u00AD]", " ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip().lower()


def main() -> None:
    if not REVIEW_FILE.exists():
        sys.exit(f"❌ {REVIEW_FILE} findes ikke - kør export_sport_review.py først.")

    wb = load_workbook(REVIEW_FILE, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        title_col = headers.index("Titel")
        image_col = headers.index("Vælg billede")
        ignore_col = headers.index("Ignorer (X)")
    except ValueError:
        sys.exit(f"❌ {REVIEW_FILE.name} har uventet kolonne-format - forventede 'Titel', 'Vælg billede' og 'Ignorer (X)'.")

    overrides = load_json(SPORT_PROGRAM_OVERRIDES_FILE, {})
    if not isinstance(overrides, dict):
        sys.exit(f"❌ {SPORT_PROGRAM_OVERRIDES_FILE.name} har uventet format (forventede et JSON-objekt).")

    updated = 0
    skipped_no_choice = 0
    unchanged = 0
    skipped_ignored = 0

    for row in ws.iter_rows(min_row=2):
        title_val = row[title_col].value
        image_val = row[image_col].value
        ignore_val = row[ignore_col].value
        if not title_val:
            continue
        if str(ignore_val or "").strip().upper() == "X":
            skipped_ignored += 1
            continue
        if not image_val or not str(image_val).strip():
            skipped_no_choice += 1
            continue

        key = normalize_title(str(title_val))
        image_name = str(image_val).strip()

        existing = overrides.get(key)
        if existing and existing.get("backdrop") == image_name and existing.get("poster") == image_name:
            unchanged += 1
            continue

        overrides[key] = {"backdrop": image_name, "poster": image_name}
        updated += 1

    save_json(SPORT_PROGRAM_OVERRIDES_FILE, overrides)

    print("=== Sport-artwork valg importeret ===")
    print(f"Fil opdateret: {SPORT_PROGRAM_OVERRIDES_FILE}")
    print(f"  Nye/ændrede overrides skrevet : {updated:,}")
    print(f"  Allerede identiske (uændret)  : {unchanged:,}")
    print(f"  Ignorerede rækker (sprunget over): {skipped_ignored:,}")
    print(f"  Rækker uden valg (sprunget over): {skipped_no_choice:,}")
    print()
    print("Kør 'python3 scripts/enrich_epg.py' igen for at få de nye billeder ind i dine XML-filer.")


if __name__ == "__main__":
    main()
