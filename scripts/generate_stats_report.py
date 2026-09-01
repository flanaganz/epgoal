#!/usr/bin/env python3
"""
generate_stats_report.py — omfattende, selvstændig HTML5-statistikrapport for
det danske backdrop-sideprojekt (danish_backdrops.py) + kanal-sundhed.

NYT (2026-08-12): bruger nu de SAMMENLAGTE kanal-grupper fra
channel_health.py ("groups", ikke "channels") og channel_priority.xlsx's
"Gruppe-nøgle (intern)"-kolonne til filtrering. Se channel_health.py og
export_channel_priority.py for baggrund om kanal-sammenlægning.

Læser:
    data/danish_artwork_cache.json
    data/danish_artwork_review.xlsx
    data/manual_artwork_overrides.xlsx
    data/danish_backdrops_run_log.json
    data/channel_health.json      (bruger "groups"-nøglen)
    data/channel_priority.xlsx    (bruger "Gruppe-nøgle (intern)"-kolonnen)

Skriver:
    output/danish_backdrops_report.html

BRUG
    python3 scripts/generate_stats_report.py
"""
from __future__ import annotations

import json
import math
import sys
import time
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"

DANISH_ARTWORK_CACHE_FILE = DATA_DIR / "danish_artwork_cache.json"
DANISH_ARTWORK_REVIEW_FILE = DATA_DIR / "danish_artwork_review.xlsx"
MANUAL_ARTWORK_OVERRIDES_FILE = DATA_DIR / "manual_artwork_overrides.xlsx"
DANISH_BACKDROPS_RUN_LOG_FILE = DATA_DIR / "danish_backdrops_run_log.json"
CHANNEL_HEALTH_FILE = DATA_DIR / "channel_health.json"
CHANNEL_PRIORITY_FILE = DATA_DIR / "channel_priority.xlsx"
REPORT_FILE = OUTPUT_DIR / "danish_backdrops_report.html"

MAX_HISTORY_ROWS = 20
MAX_UNMATCHED_ROWS = 40
MAX_MISSING_CHANNELS = 10

COLORS = {
    "found": "#22c55e", "not_found": "#334155", "approved": "#3b82f6",
    "pending": "#f59e0b", "flagged": "#ef4444", "manual": "#a855f7",
    "sport": "#ec4899", "cache_hit": "#38bdf8", "fresh_call": "#fb923c",
    "grid": "#334155",
}


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def load_review_status(path: Path) -> tuple[int, int, dict[str, str]]:
    """Returnerer (approved, flagged, notes).

    RETTET (2026-09-01): danish_artwork_review.xlsx har nu en dedikeret
    'Ignorer (X)'-kolonne (ensartet med sport_artwork_review.xlsx, se
    export_danish_artwork_review.py) i stedet for kun en fri-tekst Note der
    startede med 'Ignore'/'Ignorer'. Titler markeret 'Ignorer (X)' tælles nu
    som 'flagged' (samme "Markeret forkert"-status som før), UANSET om der
    også står en Note - og ALTID uanset 'Godkendt (X)' (samme
    sikkerhedsspærre som danish_backdrops.py's load_approved_keys()), så
    approved+flagged+pending summerer korrekt til found uden dobbeltoptælling.
    Kolonnen er valgfri for bagudkompatibilitet med ældre filer."""
    if not path.exists():
        return 0, 0, {}
    try:
        from openpyxl import load_workbook
    except ImportError:
        return 0, 0, {}

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Nøgle (intern)")
        godkendt_col = headers.index("Godkendt (X)")
        note_col = headers.index("Note") if "Note" in headers else None
    except ValueError:
        return 0, 0, {}

    ignorer_col = headers.index("Ignorer (X)") if "Ignorer (X)" in headers else None

    approved = 0
    flagged = 0
    notes: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        if not key_val:
            continue
        godkendt_val = row[godkendt_col].value
        note_val = row[note_col].value if note_col is not None else None
        ignorer_val = row[ignorer_col].value if ignorer_col is not None else None
        if str(ignorer_val or "").strip().upper() == "X":
            flagged += 1
        elif godkendt_val and str(godkendt_val).strip().upper() == "X":
            approved += 1
        elif note_val:
            flagged += 1
        if note_val:
            notes[str(key_val).strip()] = str(note_val)
    return approved, flagged, notes


def load_manual_overrides_summary(path: Path) -> list[dict]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
    except ImportError:
        return []

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Manuelle overrides"] if "Manuelle overrides" in wb.sheetnames else wb.active
    except Exception:
        return []

    headers = [c.value for c in ws[1]]
    try:
        title_col = headers.index("Titel (som i EPG)")
        channel_col = headers.index("Kanal (valgfri)")
        url_col = headers.index("Backdrop URL")
        note_col = headers.index("Note (valgfri)") if "Note (valgfri)" in headers else None
    except ValueError:
        return []

    rows = []
    for row in ws.iter_rows(min_row=2):
        title_val = row[title_col].value
        url_val = row[url_col].value
        if not title_val or not url_val:
            continue
        title = str(title_val).strip()
        if title.upper().startswith("EKSEMPEL"):
            continue
        channel_val = row[channel_col].value
        note_val = row[note_col].value if note_col is not None else None
        rows.append({
            "title": title,
            "channel": str(channel_val).strip() if channel_val else "",
            "url": str(url_val).strip(),
            "note": str(note_val).strip() if note_val else "",
        })
    return rows


def load_channel_priority(path: Path) -> set[str] | None:
    """Returnerer sæt af GRUPPE-nøgler markeret 'X', eller None hvis filen
    ikke findes (= ingen filtrering, vis alle kanaler)."""
    if not path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl er ikke installeret - kan ikke læse kanal-prioritering.", file=sys.stderr)
        return None

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Kanal-prioritering"] if "Kanal-prioritering" in wb.sheetnames else wb.active
    except Exception:
        return None

    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Gruppe-nøgle (intern)")
        follow_col = headers.index("Følg (X)")
    except ValueError:
        return None

    followed: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        follow_val = row[follow_col].value
        if key_val and follow_val and str(follow_val).strip().upper() == "X":
            followed.add(str(key_val).strip())
    return followed


def esc(s) -> str:
    return (str(s).replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
            .replace('"', "&quot;"))


def donut_chart_svg(segments: list[tuple[str, float, str]], size: int = 210, hole_ratio: float = 0.6) -> str:
    total = sum(v for _, v, _ in segments) or 1
    cx = cy = size / 2
    r_outer = size / 2 - 6
    r_inner = r_outer * hole_ratio

    def point(angle_deg, r):
        angle_rad = math.radians(angle_deg - 90)
        return cx + r * math.cos(angle_rad), cy + r * math.sin(angle_rad)

    paths = []
    angle = 0.0
    any_positive = any(v > 0 for _, v, _ in segments)
    if not any_positive:
        paths.append(f'<circle cx="{cx}" cy="{cy}" r="{(r_outer+r_inner)/2:.1f}" '
                      f'fill="none" stroke="{COLORS["not_found"]}" stroke-width="{r_outer-r_inner:.1f}"/>')
    for label, value, color in segments:
        if value <= 0:
            continue
        sweep = (value / total) * 360
        if sweep >= 359.999:
            sweep = 359.999
        end_angle = angle + sweep
        large_arc = 1 if sweep > 180 else 0
        x1o, y1o = point(angle, r_outer)
        x2o, y2o = point(end_angle, r_outer)
        x1i, y1i = point(end_angle, r_inner)
        x2i, y2i = point(angle, r_inner)
        path = (f'M {x1o:.2f},{y1o:.2f} A {r_outer:.2f},{r_outer:.2f} 0 {large_arc} 1 {x2o:.2f},{y2o:.2f} '
                f'L {x1i:.2f},{y1i:.2f} A {r_inner:.2f},{r_inner:.2f} 0 {large_arc} 0 {x2i:.2f},{y2i:.2f} Z')
        title = f"{esc(label)}: {int(value):,} ({value/total*100:.1f}%)"
        paths.append(f'<path d="{path}" fill="{color}"><title>{title}</title></path>')
        angle = end_angle

    center_text = (f'<text x="{cx}" y="{cy-6}" text-anchor="middle" class="donut-total">{int(total):,}</text>'
                   f'<text x="{cx}" y="{cy+16}" text-anchor="middle" class="donut-sub">i alt</text>')
    return f'<svg viewBox="0 0 {size} {size}" width="{size}" height="{size}">{"".join(paths)}{center_text}</svg>'


def bar_chart_svg(categories: list[str], series: list[tuple[str, list[float], str]],
                   width: int = 640, height: int = 320) -> str:
    pad_left, pad_right, pad_top, pad_bottom = 56, 20, 24, 70
    plot_w = width - pad_left - pad_right
    plot_h = height - pad_top - pad_bottom
    max_val = max((max(vals) if vals else 0) for _, vals, _ in series) or 1
    max_val = max_val * 1.15
    n_cat = len(categories) or 1
    n_series = len(series) or 1
    group_w = plot_w / n_cat
    bar_w = group_w / (n_series + 1)

    def y_scale(v):
        return pad_top + plot_h - (v / max_val) * plot_h

    parts = []
    for i in range(6):
        v = max_val / 5 * i
        y = y_scale(v)
        parts.append(f'<line x1="{pad_left}" y1="{y:.1f}" x2="{width-pad_right}" y2="{y:.1f}" '
                      f'stroke="{COLORS["grid"]}" stroke-width="1" opacity="0.4"/>')
        parts.append(f'<text x="{pad_left-10}" y="{y+4:.1f}" text-anchor="end" class="axis-label">{int(v):,}</text>')

    for cat_idx, cat in enumerate(categories):
        group_x = pad_left + cat_idx * group_w
        for s_idx, (s_name, vals, color) in enumerate(series):
            val = vals[cat_idx] if cat_idx < len(vals) else 0
            bar_x = group_x + (s_idx + 0.5) * bar_w
            bar_y = y_scale(val)
            bar_h = pad_top + plot_h - bar_y
            parts.append(f'<rect x="{bar_x:.1f}" y="{bar_y:.1f}" width="{bar_w*0.82:.1f}" height="{max(bar_h,0):.1f}" '
                         f'rx="3" fill="{color}"><title>{esc(cat)} — {esc(s_name)}: {int(val):,}</title></rect>')
        label_x = group_x + group_w / 2
        parts.append(f'<text x="{label_x:.1f}" y="{height-pad_bottom+20}" text-anchor="middle" class="axis-label" '
                     f'transform="rotate(-25 {label_x:.1f} {height-pad_bottom+20})">{esc(cat)}</text>')

    legend_x = pad_left
    legend_y = height - 18
    for s_name, _, color in series:
        parts.append(f'<rect x="{legend_x}" y="{legend_y-10}" width="12" height="12" rx="2" fill="{color}"/>')
        parts.append(f'<text x="{legend_x+18}" y="{legend_y}" class="legend-label">{esc(s_name)}</text>')
        legend_x += 18 + len(s_name) * 7 + 26

    return f'<svg viewBox="0 0 {width} {height}" width="100%" height="{height}">{"".join(parts)}</svg>'


def line_chart_svg(x_labels: list[str], series: list[tuple[str, list[float], str]],
                    width: int = 640, height: int = 260, stacked: bool = False) -> str:
    pad_left, pad_right, pad_top, pad_bottom = 56, 20, 20, 50
    plot_w = width - pad_left - pad_right
    plot_h = height - pad_top - pad_bottom
    n = len(x_labels) or 1

    if stacked:
        totals = [sum(s[1][i] if i < len(s[1]) else 0 for s in series) for i in range(n)]
        max_val = max(totals) if totals else 1
    else:
        max_val = max((max(vals) if vals else 0) for _, vals, _ in series) or 1
    max_val = max_val * 1.15 or 1

    def x_scale(i):
        return pad_left + (i / max(n - 1, 1)) * plot_w

    def y_scale(v):
        return pad_top + plot_h - (v / max_val) * plot_h

    parts = []
    for i in range(5):
        v = max_val / 4 * i
        y = y_scale(v)
        parts.append(f'<line x1="{pad_left}" y1="{y:.1f}" x2="{width-pad_right}" y2="{y:.1f}" '
                      f'stroke="{COLORS["grid"]}" stroke-width="1" opacity="0.4"/>')
        parts.append(f'<text x="{pad_left-10}" y="{y+4:.1f}" text-anchor="end" class="axis-label">{int(v):,}</text>')

    step = max(1, n // 8)
    for i, label in enumerate(x_labels):
        if i % step == 0 or i == n - 1:
            x = x_scale(i)
            parts.append(f'<text x="{x:.1f}" y="{height-pad_bottom+18}" text-anchor="middle" class="axis-label">{esc(label)}</text>')

    if stacked:
        running = [0.0] * n
        for s_name, vals, color in series:
            top = [running[i] + (vals[i] if i < len(vals) else 0) for i in range(n)]
            points_top = " ".join(f"{x_scale(i):.1f},{y_scale(top[i]):.1f}" for i in range(n))
            points_bottom = " ".join(f"{x_scale(i):.1f},{y_scale(running[i]):.1f}" for i in range(n - 1, -1, -1))
            parts.append(f'<polygon points="{points_top} {points_bottom}" fill="{color}" opacity="0.75"><title>{esc(s_name)}</title></polygon>')
            running = top
    else:
        for s_name, vals, color in series:
            points = " ".join(f"{x_scale(i):.1f},{y_scale(vals[i] if i < len(vals) else 0):.1f}" for i in range(n))
            area_points = f"{pad_left},{pad_top+plot_h} {points} {x_scale(n-1):.1f},{pad_top+plot_h}"
            parts.append(f'<polygon points="{area_points}" fill="{color}" opacity="0.12"/>')
            parts.append(f'<polyline points="{points}" fill="none" stroke="{color}" stroke-width="2.5"/>')
            for i in range(n):
                v = vals[i] if i < len(vals) else 0
                parts.append(f'<circle cx="{x_scale(i):.1f}" cy="{y_scale(v):.1f}" r="3.2" fill="{color}">'
                             f'<title>{esc(x_labels[i])} — {esc(s_name)}: {int(v):,}</title></circle>')

    legend_x = pad_left
    legend_y = 14
    for s_name, _, color in series:
        parts.append(f'<rect x="{legend_x}" y="{legend_y-9}" width="12" height="12" rx="2" fill="{color}"/>')
        parts.append(f'<text x="{legend_x+18}" y="{legend_y}" class="legend-label">{esc(s_name)}</text>')
        legend_x += 18 + len(s_name) * 7 + 26

    return f'<svg viewBox="0 0 {width} {height}" width="100%" height="{height}">{"".join(parts)}</svg>'


def pct_badge(pct: float) -> str:
    if pct >= 90:
        cls = "ok"
    elif pct >= 50:
        cls = "mid"
    else:
        cls = "warn"
    return f'<span class="badge {cls}">{pct:.1f}%</span>'


CSS = """
* { box-sizing: border-box; margin: 0; padding: 0; }
body {
    font-family: -apple-system, 'Segoe UI', Roboto, Arial, sans-serif;
    background: radial-gradient(circle at top left, #1e293b 0%, #0f172a 55%);
    color: #e2e8f0; padding: 32px 24px 64px; min-height: 100vh;
}
.wrap { max-width: 1280px; margin: 0 auto; }
header { display: flex; align-items: center; justify-content: space-between; margin-bottom: 28px; flex-wrap: wrap; gap: 12px; }
h1 { font-size: 30px; font-weight: 800; letter-spacing: -0.02em; }
h1 span.dk { color: #22c55e; } h1 span.man { color: #a855f7; }
.timestamp { color: #94a3b8; font-size: 13px; text-align: right; }
.section-title { font-size: 13px; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: #94a3b8; margin: 36px 0 14px; display:flex; align-items:center; gap:10px; }
.section-title::after { content: ""; flex:1; height:1px; background: rgba(148,163,184,0.2); }
.cards { display: grid; grid-template-columns: repeat(auto-fit, minmax(165px, 1fr)); gap: 14px; }
.card { background: rgba(30, 41, 59, 0.65); border: 1px solid rgba(148,163,184,0.15); border-radius: 14px; padding: 16px 18px; backdrop-filter: blur(6px); }
.card .label { font-size: 11px; color: #94a3b8; text-transform: uppercase; letter-spacing: 0.05em; margin-bottom: 6px; }
.card .value { font-size: 26px; font-weight: 800; }
.card .value.green { color: #22c55e; } .card .value.blue { color: #3b82f6; }
.card .value.amber { color: #f59e0b; } .card .value.purple { color: #a855f7; }
.card .value.pink { color: #ec4899; } .card .value.red { color: #ef4444; }
.grid2 { display: grid; grid-template-columns: 1fr 1fr; gap: 20px; }
.grid3 { display: grid; grid-template-columns: 1fr 1fr 1fr; gap: 20px; }
@media (max-width: 900px) { .grid2, .grid3 { grid-template-columns: 1fr; } }
.panel { background: rgba(30, 41, 59, 0.65); border: 1px solid rgba(148,163,184,0.15); border-radius: 16px; padding: 22px; margin-bottom: 22px; }
.panel h2 { font-size: 15px; font-weight: 700; margin-bottom: 16px; color: #f1f5f9; }
.panel h2 .sub { font-weight: 400; color: #94a3b8; font-size: 12px; margin-left: 8px; }
.donut-row { display: flex; align-items: center; justify-content: center; gap: 24px; flex-wrap: wrap; }
.donut-legend { display: flex; flex-direction: column; gap: 8px; }
.legend-item { display: flex; align-items: center; gap: 8px; font-size: 13px; }
.legend-dot { width: 11px; height: 11px; border-radius: 3px; flex-shrink: 0; }
.legend-item b { margin-left: auto; padding-left: 16px; }
svg text.axis-label { fill: #94a3b8; font-size: 10px; }
svg text.legend-label { fill: #cbd5e1; font-size: 11px; }
svg text.donut-total { fill: #f1f5f9; font-size: 24px; font-weight: 800; }
svg text.donut-sub { fill: #94a3b8; font-size: 11px; }
table.datatable { width: 100%; border-collapse: collapse; font-size: 12.5px; }
table.datatable th { text-align: left; color: #94a3b8; font-weight: 700; padding: 8px 10px; border-bottom: 1px solid rgba(148,163,184,0.25); position: sticky; top:0; background: rgba(15,23,42,0.95); }
table.datatable td { padding: 7px 10px; border-bottom: 1px solid rgba(148,163,184,0.08); vertical-align: top; }
table.datatable tr:hover td { background: rgba(148,163,184,0.05); }
table.datatable td.num { text-align: right; font-variant-numeric: tabular-nums; }
.table-scroll { max-height: 380px; overflow-y: auto; border-radius: 10px; }
.badge { display: inline-block; padding: 2px 9px; border-radius: 999px; font-size: 10.5px; font-weight: 700; }
.badge.warn { background: rgba(239,68,68,0.15); color: #ef4444; }
.badge.mid { background: rgba(245,158,11,0.15); color: #f59e0b; }
.badge.ok { background: rgba(34,197,94,0.15); color: #22c55e; }
.mono { font-family: 'SF Mono', Consolas, monospace; font-size: 11.5px; color: #94a3b8; word-break: break-all; }
.empty-state { color: #94a3b8; font-size: 13.5px; padding: 30px 0; text-align: center; }
.info-note { font-size: 12px; color: #94a3b8; margin-bottom: 12px; padding: 8px 12px; background: rgba(56,189,248,0.08); border-left: 3px solid #38bdf8; border-radius: 4px; }
footer { text-align: center; color: #64748b; font-size: 12px; margin-top: 36px; }
a { color: #38bdf8; }
"""


def build_html(cache: dict, run_log: list, approved: int, flagged: int, notes: dict,
               manual_rows: list[dict], channel_health: dict | None,
               channel_priority: set[str] | None) -> str:
    total_titles = len(cache)
    found = sum(1 for v in cache.values() if v.get("backdrop"))
    not_found = total_titles - found
    pending = max(found - approved - flagged, 0)
    hit_rate = (found / total_titles * 100) if total_titles else 0

    latest = run_log[-1] if run_log else None
    manual_matched_latest = latest.get("manual_titles_matched", []) if latest else []
    manual_unmatched_latest = latest.get("manual_titles_unmatched", []) if latest else []
    manual_defined_latest = latest.get("manual_defined_count", len(manual_rows)) if latest else len(manual_rows)

    last_run_str = time.strftime("%d/%m/%Y %H:%M", time.localtime(latest["timestamp"])) if latest else "—"
    generated_at = time.strftime("%d. %B %Y kl. %H:%M")

    # --- Filtrér SAMMENLAGTE kanal-grupper efter prioriteringsliste ---
    filtered_health = None
    filter_active = channel_priority is not None
    all_groups = (channel_health or {}).get("groups") or {}
    n_raw_channels = len((channel_health or {}).get("channels") or {})
    if all_groups:
        if filter_active:
            kept = {key: s for key, s in all_groups.items() if key in channel_priority}
        else:
            kept = all_groups
        total_prog = sum(s["programmes"] for s in kept.values())
        total_art = sum(s["with_artwork"] for s in kept.values())
        overall_pct = (total_art / total_prog * 100) if total_prog else 0
        filtered_health = {
            "timestamp": channel_health.get("timestamp", time.time()),
            "groups": kept,
            "overall_artwork_pct": overall_pct,
            "total_groups_before_filter": len(all_groups),
        }

    cards = [
        ("Unikke titler i cache", f"{total_titles:,}", ""),
        ("Dansk backdrop fundet", f"{found:,}", "green"),
        ("Hit-rate", f"{hit_rate:.1f}%", "purple"),
        ("Godkendt (X)", f"{approved:,}", "blue"),
        ("Afventer godkendelse", f"{pending:,}", "amber"),
        ("Manuelle overrides", f"{manual_defined_latest:,}", "pink"),
        ("Manuel matchet (seneste)", f"{len(manual_matched_latest):,}", "pink"),
        ("Kørsler i historik", f"{len(run_log):,}", ""),
    ]
    if filtered_health:
        label = "Kanal-dækning (fulgte)" if filter_active else "Kanal-artwork-dækning"
        cards.append((label, f"{filtered_health['overall_artwork_pct']:.1f}%", "green"))
    cards_html = "".join(
        f'<div class="card"><div class="label">{esc(l)}</div><div class="value {cls}">{v}</div></div>'
        for l, v, cls in cards
    )

    donut_found = donut_chart_svg([("Fundet", found, COLORS["found"]), ("Ikke fundet", not_found, COLORS["not_found"])])
    if found > 0:
        donut_review = donut_chart_svg([
            ("Godkendt", approved, COLORS["approved"]),
            ("Afventer", pending, COLORS["pending"]),
            ("Markeret forkert", flagged, COLORS["flagged"]),
        ])
    else:
        donut_review = '<div class="empty-state">Ingen fund endnu</div>'

    if latest:
        per_file = latest.get("per_file", {})
        categories = list(per_file.keys())
        sport_vals = [per_file[c].get("already_had_artwork", 0) for c in categories]
        checked_vals = [per_file[c].get("checked", 0) for c in categories]
        found_vals = [per_file[c].get("danish_found", 0) for c in categories]
        manual_vals = [per_file[c].get("manual_override_injected", 0) for c in categories]
        per_file_bar = bar_chart_svg(categories, [
            ("Sport (sprunget over)", sport_vals, COLORS["sport"]),
            ("Titler tjekket", checked_vals, COLORS["cache_hit"]),
            ("Dansk fundet", found_vals, COLORS["found"]),
            ("Manuel override", manual_vals, COLORS["manual"]),
        ])
    else:
        per_file_bar = '<div class="empty-state">Ingen kørselshistorik endnu</div>'

    if len(run_log) >= 2:
        recent = run_log[-MAX_HISTORY_ROWS:]
        x_labels = [time.strftime("%d/%m %H:%M", time.localtime(r["timestamp"])) for r in recent]
        cache_sizes = [r.get("cache_size_after", 0) for r in recent]
        unique_found_hist = [r.get("unique_found", 0) for r in recent]
        unique_pending_hist = [r.get("unique_pending", 0) for r in recent]
        manual_matched_hist = [len(r.get("manual_titles_matched", [])) for r in recent]

        cache_growth_chart = line_chart_svg(x_labels, [("Unikke titler i cache", cache_sizes, COLORS["cache_hit"])])
        found_pending_chart = line_chart_svg(x_labels, [
            ("Fundet (unikt)", unique_found_hist, COLORS["found"]),
            ("Afventer", unique_pending_hist, COLORS["pending"]),
        ])
        manual_chart = line_chart_svg(x_labels, [("Manuel override matchet", manual_matched_hist, COLORS["manual"])])

        fresh_vals = [sum(f.get("fresh_calls", 0) for f in r.get("per_file", {}).values()) for r in recent]
        cachehit_vals = [sum(f.get("cache_hits", 0) for f in r.get("per_file", {}).values()) for r in recent]
        tmdb_calls_chart = line_chart_svg(x_labels, [
            ("Cache-genbrug", cachehit_vals, COLORS["cache_hit"]),
            ("Friske TMDb-kald", fresh_vals, COLORS["fresh_call"]),
        ], stacked=True)
    else:
        empty = '<div class="empty-state">Kør scriptet flere gange for at se udvikling over tid</div>'
        cache_growth_chart = found_pending_chart = manual_chart = tmdb_calls_chart = empty

    history_rows = ""
    for r in reversed(run_log[-MAX_HISTORY_ROWS:]):
        date_str = time.strftime("%d/%m/%Y %H:%M", time.localtime(r["timestamp"]))
        history_rows += (
            f"<tr><td>{date_str}</td><td class='num'>{r.get('unique_found', 0):,}</td>"
            f"<td class='num'>{r.get('approved_count', 0):,}</td><td class='num'>{r.get('unique_pending', 0):,}</td>"
            f"<td class='num'>{len(r.get('manual_titles_matched', [])):,}</td>"
            f"<td class='num'>{r.get('cache_size_after', 0):,}</td></tr>"
        )
    history_table = (
        f'<div class="table-scroll"><table class="datatable"><thead><tr>'
        f'<th>Tidspunkt</th><th class="num">Fundet</th><th class="num">Godkendt</th><th class="num">Afventer</th>'
        f'<th class="num">Manuel matchet</th><th class="num">Cache-størrelse</th></tr></thead><tbody>{history_rows}</tbody></table></div>'
        if history_rows else '<div class="empty-state">Ingen kørselshistorik endnu</div>'
    )

    manual_def_rows = ""
    matched_set = set(manual_matched_latest)
    unmatched_set = set(manual_unmatched_latest)
    for row in sorted(manual_rows, key=lambda r: r["title"].lower()):
        title = row["title"]
        status = "ok" if title in matched_set else ("warn" if title in unmatched_set else "")
        badge = (f'<span class="badge ok">Matchet</span>' if status == "ok"
                 else f'<span class="badge warn">Ikke set i dag</span>' if status == "warn"
                 else "")
        manual_def_rows += (
            f"<tr><td>{esc(title)}</td><td>{esc(row['channel']) or '<i>alle</i>'}</td>"
            f"<td>{badge}</td><td class='mono'>{esc(row['url'][:70])}{'…' if len(row['url'])>70 else ''}</td>"
            f"<td>{esc(row['note'])}</td></tr>"
        )
    manual_def_table = (
        f'<div class="table-scroll"><table class="datatable"><thead><tr>'
        f'<th>Titel</th><th>Kanal</th><th>Status (seneste kørsel)</th><th>Backdrop URL</th><th>Note</th>'
        f'</tr></thead><tbody>{manual_def_rows}</tbody></table></div>'
        if manual_def_rows else '<div class="empty-state">Ingen manuelle overrides defineret endnu</div>'
    )

    unmatched_warning = ""
    if manual_unmatched_latest:
        items = "".join(f"<li>{esc(t)}</li>" for t in manual_unmatched_latest[:MAX_UNMATCHED_ROWS])
        unmatched_warning = (
            f'<div class="panel" style="border-color: rgba(239,68,68,0.3);">'
            f'<h2>⚠️ {len(manual_unmatched_latest)} manuel(le) override(s) ikke fundet i seneste kørsel '
            f'<span class="sub">tjek stavning, eller programmet blev bare ikke sendt i dag</span></h2>'
            f'<ul style="padding-left:20px; font-size:13px; line-height:1.9;">{items}</ul></div>'
        )

    notes_rows = ""
    if notes:
        for key, note in sorted(notes.items()):
            title = cache.get(key, {}).get("title", key)
            notes_rows += f'<tr><td>{esc(title)}</td><td>{esc(note)}</td></tr>'
    notes_table = (
        f'<div class="table-scroll"><table class="datatable"><thead><tr><th>Titel</th><th>Note</th></tr></thead>'
        f'<tbody>{notes_rows}</tbody></table></div>'
        if notes_rows else '<div class="empty-state">Ingen noter endnu</div>'
    )

    # --- Kanal-sundhed (nu på SAMMENLAGTE grupper + prioriterings-filter) ---
    channel_health_section = ""
    if filtered_health and filtered_health["groups"]:
        groups = filtered_health["groups"]
        sorted_groups = sorted(groups.items(), key=lambda kv: kv[1]["programmes"], reverse=True)

        health_rows = ""
        for key, s in sorted_groups:
            prog = s["programmes"]
            artwork_pct = (s["with_artwork"] / prog * 100) if prog else 0
            desc_pct = (s["with_desc"] / prog * 100) if prog else 0
            n_members = len(s.get("member_channel_ids", []))
            member_note = f" <span class='sub'>({n_members} kilder)</span>" if n_members > 1 else ""
            health_rows += (
                f"<tr><td>{esc(s['display_name'])}{member_note}</td><td class='num'>{prog:,}</td>"
                f"<td class='num'>{pct_badge(artwork_pct)}</td><td class='num'>{pct_badge(desc_pct)}</td></tr>"
            )
        health_table = (
            f'<div class="table-scroll"><table class="datatable"><thead><tr>'
            f'<th>Kanal</th><th class="num">Programmer</th><th class="num">Artwork %</th><th class="num">Beskrivelse %</th>'
            f'</tr></thead><tbody>{health_rows}</tbody></table></div>'
        )

        missing = [(key, s, s["programmes"] - s["with_artwork"]) for key, s in groups.items()]
        missing = [m for m in missing if m[2] > 0]
        missing.sort(key=lambda m: m[2], reverse=True)
        missing = missing[:MAX_MISSING_CHANNELS]

        missing_rows = ""
        for key, s, missing_count in missing:
            prog = s["programmes"]
            pct = (s["with_artwork"] / prog * 100) if prog else 0
            missing_rows += (
                f"<tr><td>{esc(s['display_name'])}</td><td class='num'>{missing_count:,}</td>"
                f"<td class='num'>{prog:,}</td><td class='num'>{pct_badge(pct)}</td></tr>"
            )
        missing_table = (
            f'<div class="table-scroll"><table class="datatable"><thead><tr>'
            f'<th>Kanal</th><th class="num">Manglende</th><th class="num">Total</th><th class="num">Dækning</th>'
            f'</tr></thead><tbody>{missing_rows}</tbody></table></div>'
            if missing_rows else '<div class="empty-state">Alle fulgte kanaler har 100% artwork-dækning 🎉</div>'
        )

        ch_snapshot_str = time.strftime("%d/%m/%Y %H:%M", time.localtime(filtered_health["timestamp"]))
        dedup_note = (f'<div class="info-note">🔗 {n_raw_channels:,} rå kanal-ID-varianter er automatisk '
                      f'sammenlagt til {filtered_health["total_groups_before_filter"]:,} fysiske kanaler '
                      f'(fjerner HD/FHD/Denmark-dupletter).</div>') if n_raw_channels else ""
        filter_note = ""
        if filter_active:
            n_total = filtered_health["total_groups_before_filter"]
            n_kept = len(groups)
            filter_note = (
                f'<div class="info-note">📌 Viser kun kanaler markeret "Følg (X)" i channel_priority.xlsx: '
                f'{n_kept:,} af {n_total:,} kanaler i alt.</div>'
            )
        channel_health_section = f"""
<div class="section-title">Kanal-sundhed <span style="text-transform:none;font-weight:400;color:#64748b;">(snapshot {ch_snapshot_str} · {len(groups):,} kanaler vist)</span></div>
{dedup_note}
{filter_note}
<div class="panel">
    <h2>Fulgte kanaler <span class="sub">sorteret efter flest programmer</span></h2>
    {health_table}
</div>
<div class="panel">
    <h2>Top {MAX_MISSING_CHANNELS} fulgte kanaler med mest manglende artwork <span class="sub">absolut antal, ikke procent</span></h2>
    {missing_table}
</div>
"""
    elif all_groups and filter_active:
        channel_health_section = f"""
<div class="section-title">Kanal-sundhed</div>
<div class="panel"><div class="empty-state">Ingen kanaler er markeret "Følg (X)" i channel_priority.xlsx endnu.
Åbn filen og markér de kanaler, du vil følge.</div></div>
"""
    else:
        channel_health_section = f"""
<div class="section-title">Kanal-sundhed</div>
<div class="panel"><div class="empty-state">Kør scripts/channel_health.py for at se kanal-for-kanal artwork-dækning.</div></div>
"""

    html = f"""<!DOCTYPE html>
<html lang="da">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>Danske Backdrops — Fuld Statistik</title>
<style>{CSS}</style>
</head>
<body>
<div class="wrap">
<header>
    <h1>🇩🇰 Danske <span class="dk">Backdrops</span> + <span class="man">Manuelle Overrides</span> — Statistik</h1>
    <div class="timestamp">Genereret {generated_at}<br>Seneste kørsel: {last_run_str}</div>
</header>

<div class="section-title">Overblik</div>
<div class="cards">{cards_html}</div>

<div class="section-title">TMDb-fund &amp; godkendelse</div>
<div class="grid2">
    <div class="panel">
        <h2>Dansk backdrop: fundet vs. ikke fundet</h2>
        <div class="donut-row">{donut_found}
            <div class="donut-legend">
                <div class="legend-item"><span class="legend-dot" style="background:{COLORS['found']}"></span>Fundet<b>{found:,}</b></div>
                <div class="legend-item"><span class="legend-dot" style="background:{COLORS['not_found']}"></span>Ikke fundet<b>{not_found:,}</b></div>
            </div>
        </div>
    </div>
    <div class="panel">
        <h2>Godkendelses-status <span class="sub">(af {found:,} fund)</span></h2>
        <div class="donut-row">{donut_review}
            <div class="donut-legend">
                <div class="legend-item"><span class="legend-dot" style="background:{COLORS['approved']}"></span>Godkendt<b>{approved:,}</b></div>
                <div class="legend-item"><span class="legend-dot" style="background:{COLORS['pending']}"></span>Afventer<b>{pending:,}</b></div>
                <div class="legend-item"><span class="legend-dot" style="background:{COLORS['flagged']}"></span>Markeret forkert<b>{flagged:,}</b></div>
            </div>
        </div>
    </div>
</div>

<div class="section-title">Seneste kørsel</div>
<div class="panel">
    <h2>Pr. fil <span class="sub">(sport-programmer er altid sprunget over)</span></h2>
    {per_file_bar}
</div>

<div class="section-title">Historik over tid <span style="text-transform:none;font-weight:400;color:#64748b;">(seneste {min(len(run_log), MAX_HISTORY_ROWS)} kørsler)</span></div>
<div class="grid2">
    <div class="panel"><h2>Cache-vækst</h2>{cache_growth_chart}</div>
    <div class="panel"><h2>Fundet vs. afventer</h2>{found_pending_chart}</div>
</div>
<div class="grid2">
    <div class="panel"><h2>Manuelle overrides matchet</h2>{manual_chart}</div>
    <div class="panel"><h2>TMDb-kald: cache vs. friske (stacked)</h2>{tmdb_calls_chart}</div>
</div>

<div class="panel">
    <h2>Kørselshistorik (tabel)</h2>
    {history_table}
</div>
{channel_health_section}
<div class="section-title">Manuelle overrides (film/serier, ikke sport)</div>
{unmatched_warning}
<div class="panel">
    <h2>Alle definerede manuelle overrides</h2>
    {manual_def_table}
</div>

<div class="section-title">Godkendelsesnoter</div>
<div class="panel">
    <h2>Noter fra danish_artwork_review.xlsx</h2>
    {notes_table}
</div>

<footer>epgoal · danish_backdrops + manuelle overrides + kanal-sundhed · rapporten opdateres hver gang du kører generate_stats_report.py</footer>
</div>
</body>
</html>"""
    return html


def main() -> None:
    cache = load_json(DANISH_ARTWORK_CACHE_FILE, {})
    if not cache:
        sys.exit(f"❌ {DANISH_ARTWORK_CACHE_FILE} findes ikke eller er tom - kør danish_backdrops.py først.")

    run_log = load_json(DANISH_BACKDROPS_RUN_LOG_FILE, [])
    approved, flagged, notes = load_review_status(DANISH_ARTWORK_REVIEW_FILE)
    manual_rows = load_manual_overrides_summary(MANUAL_ARTWORK_OVERRIDES_FILE)
    channel_health = load_json(CHANNEL_HEALTH_FILE, None)
    channel_priority = load_channel_priority(CHANNEL_PRIORITY_FILE)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    html = build_html(cache, run_log, approved, flagged, notes, manual_rows, channel_health, channel_priority)
    REPORT_FILE.write_text(html, encoding="utf-8")

    print("=== Fuld statistikrapport genereret ===")
    if channel_priority is not None:
        print(f"Kanal-prioritering aktiv: viser kun {len(channel_priority):,} kanaler markeret 'Følg (X)'.")
    print(f"Fil: {REPORT_FILE}")
    print(f"Åbn den i din browser (dobbeltklik filen, eller 'start {REPORT_FILE.name}' i PowerShell).")


if __name__ == "__main__":
    main()
