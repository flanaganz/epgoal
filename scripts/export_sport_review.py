#!/usr/bin/env python3
"""
export_sport_review.py — eksporterer "hit-and-miss"-sport-titler til Excel,
så du kan vælge det KORREKTE billede via en dropdown i stedet for at redigere
JSON-filer i hånden.

FORMÅL
    sport_categories.json/sport_program_overrides.json/sport_channels.json
    er tilbøjelige til at ramme forkert eller falde tilbage til et generisk
    billede for titler, der ikke matcher noget kendt mønster. Dette script
    samler ALLE de tvivlsomme tilfælde (titler der endte i kanal-fallback,
    eller som blev fanget på DR1/DR2/TV2's "partial_sport"-logik) fra:
        data/fallback_titles_log.json
        data/partial_sport_matches_log.json
    og lægger dem i ét overskueligt Excel-ark, opdelt i TRE kanalgrupper:

        Gruppe A: 3+, TV3 Sport, TV3 Max, Viaplay Sport News
        Gruppe B: TV2 Sport, TV2 Sport X, TV2
        Gruppe C: DR1, DR2

    Kolonnen "Vælg billede" er en DROPDOWN med alle .jpg/.jpeg/.png-filer der
    reelt findes i din lokale Sport/-mappe (C:\\EPGoal\\Sport) - så du aldrig
    kan vælge et filnavn, der ikke findes.

    Scriptet ÆNDRER INTET i sig selv - det er kun et eksport-værktøj. Brug
    import_sport_review.py bagefter til at skrive dine valg tilbage til
    sport_program_overrides.json.

    Kører du scriptet igen senere, BEVARES dine tidligere valg og noter
    automatisk for rækker der allerede findes i filen - kun NYE rækker
    tilføjes som friske, tomme rækker.

NYT (2026-08-27): ROBUST DROPDOWN I KOLONNE E
    Dropdown-listen bygges via et NAVNGIVET OMRÅDE ("BilledeValgListe"),
    ikke en direkte cross-sheet-formel - det er den officielt anbefalede,
    langt mere driftsikre metode i Excel, og løser tilfælde hvor dropdown'en
    tidligere kunne vise sig tom/ubrugelig. Det skjulte hjælpe-ark hedder
    "BilledeListe" (uden mellemrum/parenteser). fullCalcOnLoad er sat, så
    Excel altid genberegner ved åbning.

NYT (2026-08-28): "IGNORER"-NOTE UDELUKKER TITLEN PERMANENT FRA FREMTIDIGE
EKSPORTER
    Titler i fallback_titles_log.json / partial_sport_matches_log.json
    optrådte tidligere for ALTID i eksporten, uanset hvad du skrev i kolonne
    G ("Note") - fx skrev du "Ignorer. Benyt TMDB opslag i stedet" for
    titler som "Berlusconi: Fodbold, tv og magt", men de ville stadig dukke
    op igen, hver gang du kørte export_sport_review.py på ny, fordi
    log-filerne stadig indeholdt dem.
    Det er rettet: hvis en tidligere gemt række har en Note, der (efter
    trimning, uden hensyn til store/små bogstaver) STARTER MED "ignorer",
    bliver netop den (kanalgruppe, titel)-kombination nu UDELADT fra alle
    fremtidige eksporter - den forsvinder simpelthen fra arket, i stedet
    for at blive ved med at dukke op igen og igen. "Ignorer" gælder BEVIDST
    på tværs af ALLE kanal-varianter i gruppen (ikke kun én specifik kanal)
    - vil du ikke se en titel igen, gælder det uanset hvilken kanal-variant
    den kommer fra. Vil du fortryde, ret/slet noten (starter med "ignorer")
    i det nyeste ark, inden du kører export-scriptet igen.

RETTET (2026-08-29): NØGLE TIL AT BEVARE TIDLIGERE VALG (KANAL TILFØJET)
    "Bevar tidligere valg ved genkørsel"-logikken brugte KUN (Kanalgruppe,
    Titel) som nøgle. Men samme titel optræder ofte FLERE gange i samme
    gruppe - én række pr. kanal-variant (fx "SuperligaTilsynet" på både
    "TV3 Max.dk", "TV3 MAX HD (D) (T).dk" og "TV3 Max Denmark (DK,DA).dk").
    Da disse delte nøgle, kollapsede flere udfyldte rækker til ÉN i det
    interne opslag ved genkørsel - "sidste (ofte tomme) række vandt", og
    allerede udfyldte valg på andre rækker gik tabt.
    Nøglen til at BEVARE VALG er nu (Kanalgruppe, Kanal, Titel) - hver
    fysisk række er entydig. Bemærk at "ignorer"-nøglen (ovenfor) BEVIDST
    forbliver (Kanalgruppe, Titel) UDEN kanal, da den skal gælde bredt.
    Der laves nu også automatisk en sikkerhedskopi (sport_artwork_review.
    BACKUP.xlsx) FØR filen overskrives, og rækker der ikke længere er
    aktuelle MEN havde et udfyldt valg, samles i et ekstra ark "Forsvundne
    (havde valg)" i stedet for bare at forsvinde sporløst.

WORKFLOW
    1) python3 scripts/enrich_epg.py                  (genererer logs)
    2) python3 scripts/export_sport_review.py         (byg Excel-ark)
    3) Åbn data/sport_artwork_review.xlsx, vælg billeder via dropdown, gem
    4) python3 scripts/import_sport_review.py         (skriv valg til JSON)
    5) python3 scripts/enrich_epg.py                  (kør igen med de nye valg)

BRUG
    python3 scripts/export_sport_review.py
"""
from __future__ import annotations

import json
import shutil
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.utils import quote_sheetname

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
SPORT_DIR = ROOT / "Sport"

FALLBACK_LOG_FILE = DATA_DIR / "fallback_titles_log.json"
PARTIAL_SPORT_LOG_FILE = DATA_DIR / "partial_sport_matches_log.json"
SPORT_PROGRAM_OVERRIDES_FILE = DATA_DIR / "sport_program_overrides.json"
REVIEW_FILE = DATA_DIR / "sport_artwork_review.xlsx"
REVIEW_BACKUP_FILE = DATA_DIR / "sport_artwork_review.BACKUP.xlsx"
IGNORED_TITLES_FILE = DATA_DIR / "sport_review_ignored_titles.json"

IMAGE_EXTENSIONS = (".jpg", ".jpeg", ".png")
HELPER_SHEET_NAME = "BilledeListe"
DEFINED_NAME = "BilledeValgListe"
IGNORE_NOTE_PREFIX = "ignorer"  # case-insensitive, se docstring "IGNORER"-NOTE

# Kanal-ID'er (som de optræder i loggenes nøgler, dvs. XMLTV channel="...")
# grupperet efter dine tre ønskede kanalgrupper. Match sker som
# case-insensitive SUBSTRING, ligesom resten af pipeline'en allerede gør
# (se SportMatcher.match_channel() i enrich_epg.py) - så "TV3 Sport HD (D)
# (T).dk" matcher stadig gruppen for "tv3 sport".
CHANNEL_GROUPS = {
    "A - 3+/TV3 Sport/TV3 Max/Viaplay": ["3+", "tv3 sport", "tv3 max", "viaplay sport"],
    "B - TV2 Sport/TV2 Sport X/TV2": ["tv 2 sport x", "tv 2 sport", "tv2 sport x", "tv2 sport", "tv 2.dk", "tv2.dk", "tv 2 hd", "tv2 hd"],
    "C - DR1/DR2": ["dr1", "dr2"],
}

HEADERS = ["Kanalgruppe", "Kanal (fra log)", "Titel", "Nuværende billede", "Vælg billede", "Godkendt (X)", "Note"]


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def classify_channel(channel_id: str) -> str | None:
    """Returnerer navnet på kanalgruppen en given kanal-ID hører til, eller
    None hvis den ikke matcher nogen af de tre grupper (fx en helt anden
    kanal, der ved en fejl er endt i loggen)."""
    low = channel_id.lower()
    for group_name, patterns in CHANNEL_GROUPS.items():
        for pattern in patterns:
            if pattern in low:
                return group_name
    return None


def list_available_images(sport_dir: Path) -> list[str]:
    if not sport_dir.exists():
        print(f"⚠️  {sport_dir} findes ikke - dropdown-listen bliver tom.", file=sys.stderr)
        return []
    files = sorted(
        p.name for p in sport_dir.iterdir()
        if p.is_file() and p.suffix.lower() in IMAGE_EXTENSIONS
    )
    return files


def collect_candidates() -> dict[str, list[dict]]:
    """Samler alle tvivlsomme (kanal, titel)-kombinationer fra de to log-filer,
    grupperet efter kanalgruppe. Returnerer
    {gruppenavn: [{"channel": ..., "title": ..., "current_image": ...}, ...]}."""
    candidates: dict[str, list[dict]] = {name: [] for name in CHANNEL_GROUPS}
    seen: set[tuple[str, str, str]] = set()

    fallback_log = load_json(FALLBACK_LOG_FILE, {})
    for channel_id, titles in fallback_log.items():
        group = classify_channel(channel_id)
        if group is None:
            continue
        for title in titles:
            key = (group, channel_id, title)
            if key in seen:
                continue
            seen.add(key)
            candidates[group].append({"channel": channel_id, "title": title, "current_image": "(kanal-fallback)"})

    partial_log = load_json(PARTIAL_SPORT_LOG_FILE, {})
    for channel_id, entries in partial_log.items():
        group = classify_channel(channel_id)
        if group is None:
            continue
        for logged_key in entries:
            # Nøgleformat er "Titel -> billede_eller_TMDb_eller_null"
            if " -> " in logged_key:
                title, current_image = logged_key.rsplit(" -> ", 1)
            else:
                title, current_image = logged_key, "(ukendt)"
            key = (group, channel_id, title)
            if key in seen:
                continue
            seen.add(key)
            candidates[group].append({"channel": channel_id, "title": title, "current_image": current_image})

    for group in candidates:
        candidates[group].sort(key=lambda c: (c["title"].lower(), c["channel"].lower()))

    return candidates


def load_existing_review(path: Path) -> dict[tuple[str, str, str], dict]:
    """Returnerer {(kanalgruppe, kanal, titel): {"image": ..., "godkendt": ..., "note": ...}}
    fra en tidligere eksporteret fil, så gentagne kørsler bevarer valg.

    VIGTIGT: nøglen inkluderer KANAL (ikke kun gruppe+titel) - se docstring
    øverst i filen ("RETTET: NØGLE TIL AT BEVARE TIDLIGERE VALG") for
    baggrunden på hvorfor dette er afgørende for ikke at miste data, når
    samme titel optræder på flere kanal-varianter i samme gruppe."""
    existing: dict[tuple[str, str, str], dict] = {}
    if not path.exists():
        return existing

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        group_col = headers.index("Kanalgruppe")
        channel_col = headers.index("Kanal (fra log)")
        title_col = headers.index("Titel")
        image_col = headers.index("Vælg billede")
        godkendt_col = headers.index("Godkendt (X)")
        note_col = headers.index("Note")
    except ValueError:
        print(f"⚠️  Eksisterende {path.name} har uventet format - starter forfra uden at bevare valg.",
              file=sys.stderr)
        return existing

    for row in ws.iter_rows(min_row=2):
        group_val = row[group_col].value
        channel_val = row[channel_col].value
        title_val = row[title_col].value
        if not group_val or not title_val:
            continue
        key = (
            str(group_val).strip(),
            str(channel_val).strip() if channel_val else "",
            str(title_val).strip(),
        )
        existing[key] = {
            "image": row[image_col].value,
            "godkendt": row[godkendt_col].value,
            "note": row[note_col].value,
        }
    return existing


def is_ignored(note_value) -> bool:
    """True hvis noten (efter trim, case-insensitive) STARTER MED 'ignorer'.
    Se docstring øverst i filen ("IGNORER"-NOTE)."""
    if not note_value:
        return False
    return str(note_value).strip().lower().startswith(IGNORE_NOTE_PREFIX)


def load_ignored_titles(path: Path) -> dict[str, str]:
    """Returnerer {"kanalgruppe||titel": note} for ALLE titler der nogensinde
    er blevet markeret 'Ignorer...', UANSET om de stadig er synlige i selve
    Excel-arket. Dette er nødvendigt fordi en ignoreret titel pr. definition
    UDELADES fra arket - og uden denne separate hukommelse ville noten (og
    dermed selve "ignorer"-beslutningen) gå tabt igen, næste gang filen
    blev genindlæst, hvilket ville få titlen til fejlagtigt at dukke op
    igen ved den EFTERFØLGENDE eksport (se changelog "IGNORER"-NOTE).

    BEMÆRK: nøglen her er BEVIDST (gruppe||titel) UDEN kanal - "ignorer"
    gælder for titlen på tværs af alle kanal-varianter i gruppen. Dette er
    en anden nøgle end load_existing_review()'s (gruppe, kanal, titel), som
    bruges til at bevare selve "Vælg billede"-valget pr. fysisk række."""
    data = load_json(path, {})
    return data if isinstance(data, dict) else {}


def save_ignored_titles(path: Path, ignored: dict[str, str]) -> None:
    save_json(path, ignored)


def make_ignore_key(group_name: str, title: str) -> str:
    return f"{group_name}||{title}"


def main() -> None:
    candidates = collect_candidates()
    total_candidates = sum(len(v) for v in candidates.values())
    if total_candidates == 0:
        sys.exit(f"❌ Ingen kandidater fundet i {FALLBACK_LOG_FILE.name}/{PARTIAL_SPORT_LOG_FILE.name} "
                 "for de tre kanalgrupper - kør enrich_epg.py først.")

    available_images = list_available_images(SPORT_DIR)
    print(f"Fandt {len(available_images):,} billeder i {SPORT_DIR}")

    existing_review = load_existing_review(REVIEW_FILE)
    ignored_titles = load_ignored_titles(IGNORED_TITLES_FILE)

    # Opdater den PERMANENTE ignor-hukommelse ud fra noter i den fil, der
    # lige er blevet indlæst - dette er den ENESTE mulighed for at fange en
    # NY "ignorer"-note, før selve titlen forsvinder fra arket herunder.
    # (gruppe, kanal, titel) -> ignorer-nøgle er BEVIDST reduceret til
    # (gruppe, titel), da "ignorer" gælder på tværs af kanal-varianter.
    for (group_name, _channel, title), prior in existing_review.items():
        if is_ignored(prior.get("note")):
            ignored_titles[make_ignore_key(group_name, title)] = str(prior.get("note")).strip()

    # Sikkerhedskopi af den EKSISTERENDE fil, FØR den overskrives - så du
    # aldrig står uden en fallback, selv hvis noget uventet skulle ske.
    if REVIEW_FILE.exists():
        shutil.copy2(REVIEW_FILE, REVIEW_BACKUP_FILE)
        print(f"💾 Sikkerhedskopi af tidligere fil gemt: {REVIEW_BACKUP_FILE.name}")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sport-artwork review"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="7A1F2B")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{chr(ord('A') + len(HEADERS) - 1)}1"

    row_idx = 2
    new_count = 0
    preserved_count = 0
    ignored_count = 0
    current_keys: set[tuple[str, str, str]] = set()

    for group_name, items in candidates.items():
        for item in items:
            key = (group_name, item["channel"], item["title"])

            if make_ignore_key(group_name, item["title"]) in ignored_titles:
                ignored_count += 1
                continue

            current_keys.add(key)
            prior = existing_review.get(key)

            ws.cell(row=row_idx, column=1, value=group_name)
            ws.cell(row=row_idx, column=2, value=item["channel"])
            ws.cell(row=row_idx, column=3, value=item["title"])
            ws.cell(row=row_idx, column=4, value=item["current_image"])

            if prior and prior.get("image"):
                ws.cell(row=row_idx, column=5, value=prior["image"])
                preserved_count += 1
            else:
                new_count += 1

            if prior and prior.get("godkendt"):
                ws.cell(row=row_idx, column=6, value=prior["godkendt"])
            if prior and prior.get("note"):
                ws.cell(row=row_idx, column=7, value=prior["note"])

            row_idx += 1

    last_row = row_idx - 1

    if available_images and last_row >= 2:
        # ROBUST LØSNING: byg dropdown-listen via et NAVNGIVET OMRÅDE
        # (Excel "defined name"), ikke en direkte sheet!range-formel. Se
        # docstring øverst i filen ("ROBUST DROPDOWN I KOLONNE E").
        helper_ws = wb.create_sheet(HELPER_SHEET_NAME)
        for i, img in enumerate(available_images, start=1):
            helper_ws.cell(row=i, column=1, value=img)
        helper_ws.sheet_state = "hidden"

        defined_range = f"{quote_sheetname(HELPER_SHEET_NAME)}!$A$1:$A${len(available_images)}"
        wb.defined_names[DEFINED_NAME] = DefinedName(DEFINED_NAME, attr_text=defined_range)

        dv = DataValidation(
            type="list",
            formula1=DEFINED_NAME,
            allow_blank=True,
        )
        ws.add_data_validation(dv)
        dv.add(f"E2:E{last_row}")

    # Gennemtving fuld genberegning ved åbning, så Excel ikke risikerer at
    # vise en forældet/cached tilstand af dropdown-listen eller andre
    # afledte værdier.
    wb.calculation.fullCalcOnLoad = True

    if last_row >= 2:
        dv_godkendt = DataValidation(type="list", formula1='"X"', allow_blank=True)
        ws.add_data_validation(dv_godkendt)
        dv_godkendt.add(f"F2:F{last_row}")

    widths = {"A": 30, "B": 28, "C": 40, "D": 30, "E": 34, "F": 12, "G": 40}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    # Rækker der FANDTES i den gamle fil, men ikke længere matcher noget i de
    # aktuelle logs (OG ikke skyldes "ignorer") - vis dem tydeligt, så intet
    # forsvinder "usynligt".
    vanished_keys = set(existing_review.keys()) - current_keys
    vanished_with_data = [
        k for k in vanished_keys
        if existing_review[k].get("image") and make_ignore_key(k[0], k[2]) not in ignored_titles
    ]
    if vanished_with_data:
        vanished_ws = wb.create_sheet("Forsvundne (havde valg)")
        for col_idx, header in enumerate(["Kanalgruppe", "Kanal", "Titel", "Tidligere valgt billede", "Note"], start=1):
            c = vanished_ws.cell(row=1, column=col_idx, value=header)
            c.font = Font(bold=True)
        for r, key in enumerate(sorted(vanished_with_data), start=2):
            group_v, channel_v, title_v = key
            data = existing_review[key]
            vanished_ws.cell(row=r, column=1, value=group_v)
            vanished_ws.cell(row=r, column=2, value=channel_v)
            vanished_ws.cell(row=r, column=3, value=title_v)
            vanished_ws.cell(row=r, column=4, value=data.get("image"))
            vanished_ws.cell(row=r, column=5, value=data.get("note"))
        for col, width in {"A": 30, "B": 28, "C": 40, "D": 34, "E": 40}.items():
            vanished_ws.column_dimensions[col].width = width

    guide = wb.create_sheet("Vejledning")
    instructions = [
        ("Sådan bruger du denne fil", ""),
        ("", ""),
        ("Kolonne A - Kanalgruppe", "Angiver hvilken af dine tre kanalgrupper (A/B/C) titlen tilhører."),
        ("Kolonne D - Nuværende billede", "Viser hvad enrich_epg.py FAKTISK brugte sidste kørsel - enten et konkret filnavn, '(kanal-fallback)' hvis intet specifikt match blev fundet, eller 'TMDb' hvis et automatisk TMDb-opslag blev brugt."),
        ("Kolonne E - Vælg billede", "DROPDOWN med alle billeder der findes i din Sport/-mappe. Vælg det korrekte billede for denne titel. Virker dropdown-pilen ikke: tjek om Excel viser en gul 'Beskyttet visning'-bjælke øverst og klik 'Aktiver redigering'."),
        ("Kolonne F - Godkendt (X)", "Markér 'X' når du har bekræftet valget - bruges IKKE af import-scriptet til at afgøre om billedet skal bruges (det bruger blot om kolonne E er udfyldt), men er til DIN egen dokumentation/overblik."),
        ("Kolonne G - Note", "Din egen kommentar til de fleste formål - bruges ikke af import-scriptet. UNDTAGELSE: skriver du en note der STARTER MED 'Ignorer' (fx 'Ignorer. Benyt TMDB opslag i stedet'), vil titlen fremover blive UDELUKKET fra fremtidige eksporter for ALLE kanal-varianter i gruppen - den forsvinder simpelthen fra arket, i stedet for at blive ved med at dukke op igen."),
        ("", ""),
        ("Vigtigt", "Kør import_sport_review.py EFTER du har udfyldt og gemt denne fil, for at få dine valg skrevet til sport_program_overrides.json."),
        ("Vigtigt", "Kør derefter enrich_epg.py igen, for at få de nye billeder ind i dine XML-filer."),
        ("Vigtigt", "Kør export_sport_review.py igen når som helst - dine tidligere valg BEVARES automatisk (nøgle: gruppe+kanal+titel), kun nye rækker tilføjes. Titler markeret 'Ignorer...' i Note udelades permanent fra fremtidige eksporter (for alle kanal-varianter)."),
        ("Vigtigt", "Der laves automatisk en sikkerhedskopi (sport_artwork_review.BACKUP.xlsx) FØR filen overskrives, hver gang du kører export_sport_review.py."),
        ("Vigtigt", "Rækker der ikke længere er aktuelle, men SOM havde et udfyldt valg (og ikke er ignoreret), havner i arket 'Forsvundne (havde valg)' - intet forsvinder usporligt."),
    ]
    for r, (label, text) in enumerate(instructions, start=1):
        c1 = guide.cell(row=r, column=1, value=label)
        c1.font = Font(bold=(r == 1))
        guide.cell(row=r, column=2, value=text)
    guide.column_dimensions["A"].width = 26
    guide.column_dimensions["B"].width = 110

    wb.save(REVIEW_FILE)
    save_ignored_titles(IGNORED_TITLES_FILE, ignored_titles)

    print("=== Sport-artwork review eksporteret ===")
    print(f"Fil: {REVIEW_FILE}")
    for group_name, items in candidates.items():
        print(f"  {group_name}: {len(items):,} tvivlsomme titler i logs")
    print(f"Nye rækker denne gang: {new_count:,}  |  Tidligere valg bevaret: {preserved_count:,}"
          f"  |  Ignoreret (udelukket pga. note): {ignored_count:,}")
    if vanished_with_data:
        print(f"⚠️  {len(vanished_with_data):,} tidligere udfyldte rækker er IKKE længere aktuelle "
              f"(se ark 'Forsvundne (havde valg)') - de blev IKKE slettet nogen steder, kun ikke gentaget her.")
    print()
    print("Åbn filen, vælg billeder via dropdown i kolonne E, gem filen,")
    print("og kør derefter 'python3 scripts/import_sport_review.py'.")


if __name__ == "__main__":
    main()
