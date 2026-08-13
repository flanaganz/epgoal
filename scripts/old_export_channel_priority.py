#!/usr/bin/env python3
"""
export_channel_priority.py — eksporterer alle kanaler fundet af channel_health.py
til Excel, så du kan markere hvilke kanaler der skal FOKUSERES på i
statistikrapportens "Kanal-sundhed"-sektion. Kanaler UDEN "X" ignoreres
automatisk af generate_stats_report.py (tælles hverken i tabellen, top-10
"mest manglende"-listen, eller den overordnede dækningsprocent).

WORKFLOW
    1) python3 scripts/channel_health.py           (scanner output-filerne)
    2) python3 scripts/export_channel_priority.py  (eksporter til Excel)
    3) Åbn data/channel_priority.xlsx, markér "X" for kanaler du vil følge
    4) python3 scripts/generate_stats_report.py    (rapporten respekterer nu dit valg)

    Kør channel_health.py + export_channel_priority.py igen når som helst -
    dine eksisterende X-markeringer BEVARES automatisk for kanaler der
    allerede findes i filen. Kun NYE kanaler tilføjes som friske, tomme rækker.

BRUG
    python3 scripts/export_channel_priority.py
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"

CHANNEL_HEALTH_FILE = DATA_DIR / "channel_health.json"
CHANNEL_PRIORITY_FILE = DATA_DIR / "channel_priority.xlsx"

HEADERS = ["Kanal", "Kanal-ID (intern)", "Programmer", "Artwork %", "Følg (X)"]


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            return default
    return default


def load_existing_priority(path: Path) -> dict[str, str]:
    """Returnerer {kanal-id: 'X' eller tom} for at bevare brugerens valg ved re-eksport."""
    existing: dict[str, str] = {}
    if not path.exists():
        return existing

    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        cid_col = headers.index("Kanal-ID (intern)")
        follow_col = headers.index("Følg (X)")
    except ValueError:
        print(f"⚠️  Eksisterende {path.name} har uventet format - starter forfra uden at bevare markeringer.",
              file=sys.stderr)
        return existing

    for row in ws.iter_rows(min_row=2):
        cid_val = row[cid_col].value
        if not cid_val:
            continue
        follow_val = row[follow_col].value
        existing[str(cid_val).strip()] = str(follow_val).strip() if follow_val else ""
    return existing


def main() -> None:
    health = load_json(CHANNEL_HEALTH_FILE, None)
    if not health or not health.get("channels"):
        sys.exit(f"❌ {CHANNEL_HEALTH_FILE} findes ikke eller er tom - kør channel_health.py først.")

    channels = health["channels"]
    existing = load_existing_priority(CHANNEL_PRIORITY_FILE)
    new_count = sum(1 for cid in channels if cid not in existing)
    preserved_count = sum(1 for cid in channels if existing.get(cid) == "X")

    wb = Workbook()
    ws = wb.active
    ws.title = "Kanal-prioritering"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", start_color="0f766e")
    for col_idx, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:E1"

    sorted_channels = sorted(channels.items(), key=lambda kv: kv[1]["programmes"], reverse=True)

    for row_idx, (cid, s) in enumerate(sorted_channels, start=2):
        prog = s["programmes"]
        pct = (s["with_artwork"] / prog * 100) if prog else 0
        ws.cell(row=row_idx, column=1, value=s["display_name"])
        ws.cell(row=row_idx, column=2, value=cid)
        ws.cell(row=row_idx, column=3, value=prog)
        ws.cell(row=row_idx, column=4, value=round(pct, 1))
        prior = existing.get(cid, "")
        if prior:
            ws.cell(row=row_idx, column=5, value=prior)

    last_row = len(sorted_channels) + 1
    dv = DataValidation(type="list", formula1='"X"', allow_blank=True, showDropDown=False)
    ws.add_data_validation(dv)
    dv.add(f"E2:E{last_row}")

    widths = {"A": 38, "B": 42, "C": 14, "D": 12, "E": 10}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    ws2 = wb.create_sheet("Vejledning")
    instructions = [
        ("Sådan bruger du denne fil", ""),
        ("", ""),
        ("Kolonne E - Følg (X)", "Markér 'X' for de kanaler du VIL følge i statistikrapporten. Kanaler UDEN 'X' bliver automatisk IGNORERET i rapportens 'Kanal-sundhed'-sektion (tælles hverken i tabellen, top-10-listen, eller den samlede dækningsprocent)."),
        ("", ""),
        ("Vigtigt", "Kør export_channel_priority.py igen, når nye kanaler dukker op i channel_health.json - dine eksisterende X-markeringer bevares automatisk, kun nye kanaler tilføjes."),
        ("Vigtigt", "Kør generate_stats_report.py efter du har gemt denne fil, for at få rapporten til at respektere dit valg."),
        ("Vigtigt", "Kolonnerne 'Programmer' og 'Artwork %' er et SNAPSHOT fra sidste channel_health.py-kørsel - de opdateres ikke automatisk her, men bruges kun til at hjælpe dig vælge."),
    ]
    for row_idx, (label, text) in enumerate(instructions, start=1):
        c1 = ws2.cell(row=row_idx, column=1, value=label)
        c1.font = Font(bold=(row_idx == 1))
        ws2.cell(row=row_idx, column=2, value=text)
    ws2.column_dimensions['A'].width = 24
    ws2.column_dimensions['B'].width = 100

    wb.save(CHANNEL_PRIORITY_FILE)

    print("=== Kanal-prioriteringsliste eksporteret ===")
    print(f"Fil: {CHANNEL_PRIORITY_FILE}")
    print(f"Kanaler i alt: {len(sorted_channels):,}")
    print(f"  - Nye kanaler tilføjet denne gang : {new_count:,}")
    print(f"  - Eksisterende X-markeringer bevaret: {preserved_count:,}")
    print()
    print("Åbn filen, markér 'X' for de kanaler du vil følge, gem filen,")
    print("og kør derefter 'python3 scripts/generate_stats_report.py' igen.")


if __name__ == "__main__":
    main()
