#!/usr/bin/env python3
"""
danish_backdrops.py — separat sideprojekt til danske TMDb-backdrops (Mac mini / always-on)

Se README/tidligere dokumentation for fuld baggrund. Denne version tilføjer:
- Detaljeret logning af manuelle overrides (matchet/ikke-matchet titel-lister,
  ikke kun optællinger) til data/danish_backdrops_run_log.json, så
  generate_stats_report.py kan vise historik og "mistænkte tastefejl"-lister.

NB: Dette script håndterer KUN backdrops (16:9). Posters er fjernet fra
hele pipelinen, da UHF ikke bruger dem.

NYT (2026-08-24): DIFFERENTIERET CACHE-LEVETID
    Tidligere blev BÅDE "fandt et dansk backdrop" og "fandt intet" cachet med
    samme levetid (cache_max_age_days, typisk 30 dage). Det betød, at hvis
    TMDb IKKE havde et dansk backdrop ved første tjek, men et senere blev
    tilføjet (fx af brugeren selv på tmdb.com), ville scriptet ikke opdage
    det nye billede i op til 30 dage - fordi "ikke fundet"-resultatet blev
    genbrugt fra cachen uden et nyt TMDb-opslag.

    Det er rettet: "fandt et backdrop"-resultater beholder den fulde
    cache_max_age_days-levetid (stabilt, ændrer sig praktisk talt aldrig).
    "Ikke fundet"-resultater udløber nu meget hurtigere
    (NOT_FOUND_CACHE_MAX_AGE_DAYS, default 2 dage), så scriptet automatisk
    prøver igen snart efter et nyt dansk billede er blevet tilføjet på TMDb,
    uden at du behøver slette/redigere cachen manuelt.
"""
from __future__ import annotations

import argparse
import difflib
import json
import os
import re
import subprocess
import sys
import time
import unicodedata
from pathlib import Path
from xml.etree import ElementTree as ET

import requests

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT_DIR = ROOT / "output"
CONFIG_FILE = ROOT / "config.json"

DANISH_ARTWORK_CACHE_FILE = DATA_DIR / "danish_artwork_cache.json"
DANISH_ARTWORK_REVIEW_FILE = DATA_DIR / "danish_artwork_review.xlsx"
DANISH_BACKDROPS_RUN_LOG_FILE = DATA_DIR / "danish_backdrops_run_log.json"
MANUAL_ARTWORK_OVERRIDES_FILE = DATA_DIR / "manual_artwork_overrides.xlsx"
MAX_RUN_LOG_ENTRIES = 200

# NYT: hvor længe et "ikke fundet"-resultat er gyldigt, FØR scriptet prøver
# TMDb igen. Sat markant lavere end cache_max_age_days (for FUNDNE billeder),
# da manglende danske backdrops er langt mere sandsynlige at ændre sig over
# tid (nye brugerbidrag på TMDb) end allerede fundne, stabile billeder.
NOT_FOUND_CACHE_MAX_AGE_DAYS = 10

ENV_FILE = ROOT / ".env"
if ENV_FILE.exists():
    for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        os.environ.setdefault(k.strip(), v.strip().strip('"').strip("'"))

TMDB_API_KEY = os.environ.get("TMDB_API_KEY")
TMDB_BASE = "https://api.themoviedb.org/3"
IMAGE_BASE = "https://image.tmdb.org/t/p"

MATCH_SIMILARITY_MIN = 0.55
REQUEST_SLEEP_SECONDS = 0.05

INVISIBLE_CHARS_PATTERN = re.compile(r"[\u200B\u200C\u200D\u2060\uFEFF\u00AD]")

SESSION = requests.Session()


def load_json(path: Path, default):
    if path.exists():
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            print(f"⚠️  Kunne ikke læse {path}, bruger default", file=sys.stderr)
    return default


def save_json(path: Path, data) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2, sort_keys=True), encoding="utf-8")


def normalize_title(title: str) -> str:
    t = title.strip()
    t = unicodedata.normalize("NFKC", t)
    t = INVISIBLE_CHARS_PATTERN.sub(" ", t)
    t = re.sub(r"\s+", " ", t)
    return t.strip().lower()


def load_manual_overrides(path: Path) -> tuple[dict[str, list[dict]], dict[str, str]]:
    """Returnerer (index, display_titles).
    index: normaliseret titel -> liste af {'channel', 'backdrop_url'}.
    display_titles: normaliseret titel -> original (pænt formateret) titel,
    til brug i rapportering."""
    index: dict[str, list[dict]] = {}
    display_titles: dict[str, str] = {}
    if not path.exists():
        return index, display_titles

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl er ikke installeret - kan ikke læse manuelle overrides. "
              "Kør: pip install openpyxl", file=sys.stderr)
        return index, display_titles

    try:
        wb = load_workbook(path, data_only=True)
        ws = wb["Manuelle overrides"] if "Manuelle overrides" in wb.sheetnames else wb.active
    except Exception as exc:
        print(f"⚠️  Kunne ikke åbne {path.name}: {exc}", file=sys.stderr)
        return index, display_titles

    headers = [c.value for c in ws[1]]
    try:
        title_col = headers.index("Titel (som i EPG)")
        channel_col = headers.index("Kanal (valgfri)")
        url_col = headers.index("Backdrop URL")
    except ValueError:
        print(f"⚠️  {path.name} mangler forventede kolonner - ingen manuelle overrides indlæst.", file=sys.stderr)
        return index, display_titles

    for row in ws.iter_rows(min_row=2):
        title_val = row[title_col].value
        url_val = row[url_col].value
        if not title_val or not url_val:
            continue
        title = str(title_val).strip()
        url = str(url_val).strip()
        if not title or not url or title.upper().startswith("EKSEMPEL"):
            continue
        norm = normalize_title(title)
        display_titles[norm] = title
        channel_val = row[channel_col].value
        channel = str(channel_val).strip().lower() if channel_val else ""
        index.setdefault(norm, []).append({"channel": channel, "backdrop_url": url})
    return index, display_titles


def resolve_manual_override(title: str, channel_id: str, manual_index: dict[str, list[dict]]) -> str | None:
    norm = normalize_title(title)
    entries = manual_index.get(norm)
    if not entries:
        return None
    channel_low = (channel_id or "").lower()
    for e in entries:
        if e["channel"] and e["channel"] in channel_low:
            return e["backdrop_url"]
    for e in entries:
        if not e["channel"]:
            return e["backdrop_url"]
    return None


def load_approved_keys(review_path: Path) -> set[str] | None:
    if not review_path.exists():
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("⚠️  openpyxl er ikke installeret.", file=sys.stderr)
        return None

    wb = load_workbook(review_path, data_only=True)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    try:
        key_col = headers.index("Nøgle (intern)")
        godkendt_col = headers.index("Godkendt (X)")
    except ValueError:
        return set()

    approved: set[str] = set()
    for row in ws.iter_rows(min_row=2):
        key_val = row[key_col].value
        godkendt_val = row[godkendt_col].value
        if key_val and str(godkendt_val).strip().upper() == "X":
            approved.add(str(key_val).strip())
    return approved


def tmdb_search(title: str) -> tuple[str, int] | None:
    resp = SESSION.get(
        f"{TMDB_BASE}/search/multi",
        params={"api_key": TMDB_API_KEY, "query": title, "language": "da-DK", "include_adult": "false"},
        timeout=15,
    )
    resp.raise_for_status()
    results = [r for r in resp.json().get("results", []) if r.get("media_type") in ("tv", "movie")]
    if not results:
        return None

    def score(r):
        name = r.get("name") or r.get("title") or ""
        similarity = difflib.SequenceMatcher(None, name.lower(), title.lower()).ratio()
        popularity_bonus = min(r.get("popularity", 0), 50) / 50
        return similarity * 2 + popularity_bonus

    results.sort(key=score, reverse=True)
    best = results[0]
    best_name = best.get("name") or best.get("title") or ""
    similarity = difflib.SequenceMatcher(None, best_name.lower(), title.lower()).ratio()
    if similarity < MATCH_SIMILARITY_MIN:
        return None
    return best["media_type"], best["id"]


def tmdb_danish_backdrop(media_type: str, tmdb_id: int) -> str | None:
    """Henter bedste danske backdrop for et TMDb-objekt. Posters håndteres ikke."""
    resp = SESSION.get(
        f"{TMDB_BASE}/{media_type}/{tmdb_id}/images",
        params={"api_key": TMDB_API_KEY, "include_image_language": "da"},
        timeout=15,
    )
    resp.raise_for_status()
    data = resp.json()

    backdrops = data.get("backdrops", [])
    if not backdrops:
        return None
    backdrops = sorted(backdrops, key=lambda b: -b.get("vote_average", 0))
    return backdrops[0]["file_path"]


def resolve_danish_artwork(raw_title: str, cache: dict, cache_max_age_days: int,
                            not_found_cache_max_age_days: int,
                            backdrop_size: str) -> tuple[str | None, bool]:
    """Returnerer (backdrop_url eller None, from_cache).

    NYT: bruger differentieret levetid. Et cachet FUND (backdrop != None)
    bruges op til cache_max_age_days gammel. Et cachet "IKKE fundet"
    (backdrop == None) bruges kun op til not_found_cache_max_age_days gammel
    - herefter forsøges et nyt, friskt TMDb-opslag, så nyligt tilføjede
    danske billeder på TMDb bliver opdaget langt hurtigere."""
    key = normalize_title(raw_title)
    cached = cache.get(key)
    if cached is not None:
        age_days = (time.time() - cached.get("ts", 0)) / 86400
        had_backdrop = bool(cached.get("backdrop"))
        max_age = cache_max_age_days if had_backdrop else not_found_cache_max_age_days
        if age_days < max_age:
            return cached.get("backdrop"), True

    backdrop_url = None
    try:
        match = tmdb_search(raw_title)
        if match:
            media_type, tmdb_id = match
            b_path = tmdb_danish_backdrop(media_type, tmdb_id)
            if b_path:
                backdrop_url = f"{IMAGE_BASE}/{backdrop_size}{b_path}"
    except requests.RequestException as exc:
        print(f"   TMDb-fejl for '{raw_title}': {exc}", file=sys.stderr)
        return None, False

    cache[key] = {"title": raw_title, "backdrop": backdrop_url, "ts": time.time()}
    return backdrop_url, False


def process_xml_file(xml_path: Path, cache: dict, cache_max_age_days: int,
                      not_found_cache_max_age_days: int,
                      backdrop_size: str, limit: int | None,
                      titles_processed_this_run: set, approved_keys: set[str] | None,
                      all_found_keys: set[str], manual_index: dict[str, list[dict]],
                      manual_titles_matched: set[str]) -> dict:
    tree = ET.parse(xml_path)
    root = tree.getroot()

    stats = {
        "programmes": 0, "already_had_artwork": 0, "checked": 0,
        "danish_found": 0, "danish_not_found": 0, "danish_injected": 0,
        "cache_hits": 0, "fresh_calls": 0, "manual_override_injected": 0,
        "rechecked_after_not_found": 0,
    }
    resolved_this_file: dict[str, str | None] = {}

    for programme in root.findall("programme"):
        stats["programmes"] += 1

        if programme.find("icon") is not None or programme.find("backdrop") is not None:
            stats["already_had_artwork"] += 1
            continue

        title_el = programme.find("title")
        if title_el is None or not title_el.text:
            continue
        title = title_el.text

        chan_id = programme.get("channel", "")
        manual_url = resolve_manual_override(title, chan_id, manual_index)
        if manual_url:
            el = ET.SubElement(programme, "icon")
            el.set("src", manual_url)
            stats["manual_override_injected"] += 1
            manual_titles_matched.add(normalize_title(title))
            continue

        norm = normalize_title(title)

        if limit is not None and norm not in titles_processed_this_run and len(titles_processed_this_run) >= limit:
            continue

        if norm not in resolved_this_file:
            was_cached_not_found = (
                norm in cache
                and not cache[norm].get("backdrop")
                and (time.time() - cache[norm].get("ts", 0)) / 86400 >= not_found_cache_max_age_days
                and (time.time() - cache[norm].get("ts", 0)) / 86400 < cache_max_age_days
            )
            backdrop_url, from_cache = resolve_danish_artwork(
                title, cache, cache_max_age_days, not_found_cache_max_age_days, backdrop_size
            )
            resolved_this_file[norm] = backdrop_url
            titles_processed_this_run.add(norm)
            stats["checked"] += 1
            if from_cache:
                stats["cache_hits"] += 1
            else:
                stats["fresh_calls"] += 1
                if was_cached_not_found:
                    stats["rechecked_after_not_found"] += 1
                time.sleep(REQUEST_SLEEP_SECONDS)
            if backdrop_url:
                stats["danish_found"] += 1
                all_found_keys.add(norm)
            else:
                stats["danish_not_found"] += 1

        backdrop_url = resolved_this_file[norm]
        if backdrop_url and approved_keys is not None and norm in approved_keys:
            el = ET.SubElement(programme, "icon")
            el.set("src", backdrop_url)
            stats["danish_injected"] += 1

    tree.write(xml_path, encoding="utf-8", xml_declaration=True)
    return stats


def append_run_log(log_path: Path, per_file_stats: dict, grand_total: dict,
                    cache_size_before: int, cache_size_after: int,
                    approved_count: int, review_exists: bool,
                    unique_found: int, unique_pending: int,
                    manual_defined_count: int,
                    manual_titles_matched_display: list[str],
                    manual_titles_unmatched_display: list[str]) -> None:
    history = load_json(log_path, [])
    if not isinstance(history, list):
        history = []

    history.append({
        "timestamp": time.time(),
        "date_str": time.strftime("%Y-%m-%d %H:%M:%S"),
        "per_file": per_file_stats,
        "totals": grand_total,
        "cache_size_before": cache_size_before,
        "cache_size_after": cache_size_after,
        "approved_count": approved_count,
        "unique_found": unique_found,
        "unique_pending": unique_pending,
        "review_file_existed": review_exists,
        "manual_defined_count": manual_defined_count,
        "manual_titles_matched": manual_titles_matched_display,
        "manual_titles_unmatched": manual_titles_unmatched_display,
    })
    history = history[-MAX_RUN_LOG_ENTRIES:]
    save_json(log_path, history)


def git_push(repo_dir: Path, commit_message: str) -> None:
    print("\n⬆️  Committer og pusher til GitHub ...")
    try:
        subprocess.run(["git", "add", "-A"], cwd=repo_dir, check=False)
        result = subprocess.run(
            ["git", "commit", "-m", commit_message], cwd=repo_dir, check=False, capture_output=True, text=True
        )
        if "nothing to commit" in (result.stdout + result.stderr).lower():
            print("   Ingen ændringer at committe.")
            return

        subprocess.run(["git", "fetch", "origin"], cwd=repo_dir, check=False,
                        capture_output=True, text=True)
        push_result = subprocess.run(
            ["git", "push", "origin", "main"], cwd=repo_dir, check=False, capture_output=True, text=True
        )
        if push_result.returncode == 0:
            print("✅ Git push lykkedes.")
            return

        combined_output = (push_result.stdout + push_result.stderr).lower()
        if "rejected" in combined_output or "fetch first" in combined_output or "non-fast-forward" in combined_output:
            print("⚠️  Push afvist (fjernrepo har nyere commits) - forsøger 'git pull --rebase' og prøver igen ...")
            rebase_result = subprocess.run(
                ["git", "pull", "--rebase", "origin", "main"], cwd=repo_dir, check=False,
                capture_output=True, text=True,
            )
            if rebase_result.returncode != 0:
                print("❌ 'git pull --rebase' fejlede - løs konflikten manuelt:", file=sys.stderr)
                print(rebase_result.stdout + rebase_result.stderr, file=sys.stderr)
                return
            retry_result = subprocess.run(
                ["git", "push", "origin", "main"], cwd=repo_dir, check=False, capture_output=True, text=True
            )
            if retry_result.returncode == 0:
                print("✅ Git push lykkedes efter rebase.")
            else:
                print("❌ Git push fejlede STADIG efter rebase - tjek manuelt:", file=sys.stderr)
                print(retry_result.stdout + retry_result.stderr, file=sys.stderr)
        else:
            print("❌ Git push fejlede af en anden årsag:", file=sys.stderr)
            print(push_result.stdout + push_result.stderr, file=sys.stderr)
    except FileNotFoundError:
        print("⚠️  git blev ikke fundet i PATH — spring commit/push over.", file=sys.stderr)


def main() -> None:
    parser = argparse.ArgumentParser(description="Tilføj GODKENDTE danske TMDb-backdrops (som <icon>) til ikke-sport-programmer.")
    parser.add_argument("--limit", type=int, default=None)
    parser.add_argument("--files", nargs="*", default=None)
    args = parser.parse_args()

    if not TMDB_API_KEY:
        sys.exit("❌ TMDB_API_KEY er ikke sat i .env.")

    config = load_json(CONFIG_FILE, {})
    sources = config.get("sources", [])
    if not sources:
        sys.exit("❌ Ingen kilder defineret i config.json.")

    source_names = [s["name"] for s in sources]
    if args.files:
        source_names = [n for n in source_names if n in args.files]
        if not source_names:
            sys.exit("❌ Ingen af de angivne --files matcher kilderne i config.json.")

    backdrop_size = config.get("image", {}).get("backdrop_size", "w1280")
    cache_max_age_days = config.get("cache_max_age_days", 30)
    not_found_cache_max_age_days = config.get("not_found_cache_max_age_days", NOT_FOUND_CACHE_MAX_AGE_DAYS)
    git_cfg = config.get("git", {})

    DATA_DIR.mkdir(parents=True, exist_ok=True)

    cache = load_json(DANISH_ARTWORK_CACHE_FILE, {})
    cache_size_before = len(cache)

    approved_keys = load_approved_keys(DANISH_ARTWORK_REVIEW_FILE)
    review_exists = DANISH_ARTWORK_REVIEW_FILE.exists()

    manual_index, manual_display_titles = load_manual_overrides(MANUAL_ARTWORK_OVERRIDES_FILE)
    manual_titles_matched: set[str] = set()

    print("=== Danske TMDb-backdrops (separat sideprojekt) — skrives som <icon> ===")
    print(f"Cache indeholder {cache_size_before:,} tidligere opslag (levetid: {cache_max_age_days} dage for fund, "
          f"{not_found_cache_max_age_days} dage for 'ikke fundet')")
    if MANUAL_ARTWORK_OVERRIDES_FILE.exists():
        print(f"Manuelle overrides indlæst: {len(manual_index):,} unikke titler")
    else:
        print(f"ℹ️  {MANUAL_ARTWORK_OVERRIDES_FILE.name} findes ikke - ingen manuelle overrides denne gang.")

    if approved_keys is None:
        print(f"⚠️  {DANISH_ARTWORK_REVIEW_FILE.name} findes IKKE endnu. Ingen TMDb-fund injiceres denne gang.")
        approved_keys = set()
    else:
        print(f"Godkendelsesfil fundet: {len(approved_keys):,} unikke titler markeret med X.")

    if args.limit:
        print(f"⚠️  TEST-TILSTAND: maks {args.limit} nye unikke titler slås op.")
    print(f"Behandler filer: {', '.join(source_names)}")

    titles_processed_this_run: set = set()
    all_found_keys: set[str] = set()
    grand_total = {
        "programmes": 0, "already_had_artwork": 0, "checked": 0,
        "danish_found": 0, "danish_not_found": 0, "danish_injected": 0,
        "cache_hits": 0, "fresh_calls": 0, "manual_override_injected": 0,
        "rechecked_after_not_found": 0,
    }
    per_file_stats: dict[str, dict] = {}

    for name in source_names:
        xml_path = OUTPUT_DIR / f"{name}.xml"
        if not xml_path.exists():
            print(f"\n⚠️  {xml_path} findes ikke - kør enrich_epg.py først. Springer over.")
            continue

        print(f"\n📄 Behandler {xml_path.name} ...")
        stats = process_xml_file(
            xml_path, cache, cache_max_age_days, not_found_cache_max_age_days, backdrop_size,
            args.limit, titles_processed_this_run, approved_keys, all_found_keys,
            manual_index, manual_titles_matched,
        )
        save_json(DANISH_ARTWORK_CACHE_FILE, cache)

        print(f"   Programmer i alt: {stats['programmes']:,} | Sprunget over (sport): {stats['already_had_artwork']:,} "
              f"| Manuel override: {stats['manual_override_injected']:,} | Tjekket: {stats['checked']:,} "
              f"| Dansk fundet: {stats['danish_found']:,} | Godkendt+indsat: {stats['danish_injected']:,}")
        if stats["rechecked_after_not_found"]:
            print(f"   🔄 Gen-tjekket efter tidligere 'ikke fundet': {stats['rechecked_after_not_found']:,}")

        per_file_stats[name] = stats
        for k in grand_total:
            grand_total[k] += stats[k]

    save_json(DANISH_ARTWORK_CACHE_FILE, cache)
    cache_size_after = len(cache)

    unique_found = len(all_found_keys)
    unique_approved_and_found = len(all_found_keys & approved_keys)
    unique_pending = unique_found - unique_approved_and_found

    all_manual_keys = set(manual_index.keys())
    unmatched_manual_keys = all_manual_keys - manual_titles_matched
    manual_matched_display = sorted(manual_display_titles.get(k, k) for k in manual_titles_matched)
    manual_unmatched_display = sorted(manual_display_titles.get(k, k) for k in unmatched_manual_keys)

    append_run_log(
        DANISH_BACKDROPS_RUN_LOG_FILE, per_file_stats, grand_total,
        cache_size_before, cache_size_after, len(approved_keys), review_exists,
        unique_found, unique_pending,
        len(manual_index), manual_matched_display, manual_unmatched_display,
    )

    print("\n📊 SAMLET RAPPORT")
    print("--------------------------------")
    print(f"Programmer i alt              : {grand_total['programmes']:,}")
    print(f"Sprunget over (sport)          : {grand_total['already_had_artwork']:,}")
    print(f"Manuelle overrides indsat      : {grand_total['manual_override_injected']:,} "
          f"({len(manual_titles_matched):,} unikke ud af {len(manual_index):,} defineret)")
    if manual_unmatched_display:
        print(f"⚠️  {len(manual_unmatched_display):,} manuelle overrides IKKE brugt (tjek stavning):")
        for t in manual_unmatched_display:
            print(f"     - {t}")
    print(f"Unikke titler med dansk backdrop: {unique_found:,} (godkendt: {unique_approved_and_found:,}, "
          f"afventer: {unique_pending:,})")
    if grand_total["rechecked_after_not_found"]:
        print(f"🔄 Titler gen-tjekket efter tidligere 'ikke fundet' (kan nu være fundet): "
              f"{grand_total['rechecked_after_not_found']:,}")
    print(f"Cache voksede fra {cache_size_before:,} til {cache_size_after:,}")
    print("--------------------------------")

    if git_cfg.get("enabled", True):
        prefix = "Auto-sync EPG (danske backdrops)"
        git_push(ROOT, f"{prefix} {time.strftime('%Y-%m-%d %H:%M:%S')}")

    print("\n=== Færdig. ===")


if __name__ == "__main__":
    main()
